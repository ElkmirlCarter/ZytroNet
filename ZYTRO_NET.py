#!/usr/bin/env python3
"""
ZytroNet Monitoring Engine v3.0
Real-time network monitoring with live web dashboard.
Built by Elkmirl Kuuku Carter.
"""

import os
import sys
import re
import time
import json
import atexit
import base64
import hmac
import socket
import platform
import threading
import argparse
import statistics
import ipaddress
import subprocess
import concurrent.futures
from datetime import datetime
from collections import deque

# Global shutdown coordination (shared by all long-running threads)
_shutdown_event = threading.Event()

# Lock-rank foundation (feature behavior toggling uses this later)
BEEP_ENABLED: bool = True
_beep_lock: threading.Lock = threading.Lock()
_ping_executor = concurrent.futures.ThreadPoolExecutor(max_workers=4)
_ping_in_flight = set()
_ping_in_flight_lock = threading.Lock()

# ── Optional imports ──────────────────────────────────────────────────────────
try:
    import psutil
    PSUTIL_OK = True
except ImportError:
    PSUTIL_OK = False

try:
    from flask import Flask, Response, render_template, render_template_string
    FLASK_OK = True
except ImportError:
    FLASK_OK = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Platform ──────────────────────────────────────────────────────────────────
IS_WIN = platform.system() == "Windows"
IS_LIN = platform.system() == "Linux"
IS_MAC = platform.system() == "Darwin"

# ── ANSI ──────────────────────────────────────────────────────────────────────
class A:
    RST  = "\033[0m";  BOLD = "\033[1m"
    RED  = "\033[91m"; GRN  = "\033[92m"; YLW  = "\033[93m"
    CYN  = "\033[96m"; WHT  = "\033[97m"; GRY  = "\033[90m"
    MGT  = "\033[95m"
    CLR  = "\033[2J\033[H"
    HIDE = "\033[?25l"
    SHOW = "\033[?25h"

NO_COLOR = False

def cl(text, *codes):
    if NO_COLOR: return str(text)
    return "".join(codes) + str(text) + A.RST

def status_col(s):
    return {"ONLINE": A.GRN, "DEGRADED": A.YLW,
            "ISP_FAILURE": A.YLW, "OFFLINE": A.RED}.get(s, A.GRY)

# ── Argument Parser ───────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser(description="ZytroNet v3.0 — Real-Time Network Monitor")
    p.add_argument("--interval",     type=float, default=3.0,
                   help="Refresh interval in seconds (default: 3)")
    p.add_argument("--targets",      nargs="+",  default=["8.8.8.8", "1.1.1.1"],
                   help="Ping targets (default: 8.8.8.8 1.1.1.1)")
    p.add_argument("--ping-count",   type=int,   default=3,
                   help="Pings per target per cycle (default: 3)")
    p.add_argument("--report",       type=str,   default="zytronet_report.html",
                   help="HTML report output path (default: zytronet_report.html)")
    p.add_argument("--web-port",     type=int,   default=5000,
                   help="Live web dashboard port (default: 5000)")
    p.add_argument("--no-arp",       action="store_true",
                   help="Disable ARP device scanning")
    p.add_argument("--no-web",       action="store_true",
                   help="Disable live web dashboard")
    p.add_argument("--no-color",     action="store_true",
                   help="Disable terminal colors")
    p.add_argument("--arp-interval", type=float, default=60.0,
                   help="ARP scan interval in seconds (default: 60)")
    p.add_argument("--history",      type=int,   default=50,
                   help="Max metric samples in memory (default: 50)")
    p.add_argument("--max-age",      type=float, default=120,
                   help="Max age of data in seconds before clearing (default: 120)")
    p.add_argument("--web-password", type=str, default=None,
                   help="Optional web dashboard basic auth password")
    return p.parse_args()

# ── Privilege Check ───────────────────────────────────────────────────────────
def is_root():
    if IS_WIN:
        try:
            import ctypes
            return ctypes.windll.shell32.IsUserAnAdmin() != 0
        except Exception:
            return False
    return os.geteuid() == 0

# ── Utilities ─────────────────────────────────────────────────────────────────
def fmt_bytes(b):
    for u in ["B", "KB", "MB", "GB", "TB"]:
        if abs(b) < 1024: return f"{b:.1f} {u}"
        b /= 1024
    return f"{b:.1f} PB"

def fmt_speed(bps): return fmt_bytes(bps) + "/s"

def fmt_dur(s):
    h, m, sec = int(s // 3600), int((s % 3600) // 60), int(s % 60)
    if h:  return f"{h}h {m}m {sec}s"
    if m:  return f"{m}m {sec}s"
    return f"{sec}s"

def sparkline(vals):
    bars = " ▁▂▃▄▅▆▇█"
    if not vals: return cl("no data", A.GRY)
    mn, mx = min(vals), max(vals)
    rng = mx - mn if mx != mn else 1
    out = ""
    for v in vals:
        idx = int((v - mn) / rng * (len(bars) - 1))
        col = A.GRN if v < 60 else A.YLW if v < 150 else A.RED
        out += cl(bars[idx], col)
    return out

def quality_score(latency, loss, jitter):
    if latency is None: return 0
    s = (max(0, 100 - latency / 3) * 0.50 +
         max(0, 100 - loss * 4)    * 0.35 +
         max(0, 100 - jitter * 2)  * 0.15)
    return round(max(0, min(100, s)), 1)

def score_label(score):
    if score >= 85: return "Excellent"
    if score >= 70: return "Good"
    if score >= 50: return "Fair"
    if score >= 25: return "Poor"
    return "Critical"

def score_color_css(score):
    if score >= 85: return "#00ff88"
    if score >= 70: return "#84cc16"
    if score >= 50: return "#ffd700"
    if score >= 25: return "#ff8c00"
    return "#ff4444"

def beep():
    with _beep_lock:
        _val = BEEP_ENABLED
    if not _val:
        return
    try:
        if IS_WIN:
            import winsound
            winsound.Beep(1000, 300)
            winsound.Beep(1200, 300)
        else:
            sys.stdout.write("\a\a")
            sys.stdout.flush()
    except: pass

# ── Network Info ──────────────────────────────────────────────────────────────
def get_primary_interface():
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
        if PSUTIL_OK:
            for iface, addrs in psutil.net_if_addrs().items():
                for a in addrs:
                    if a.family == socket.AF_INET and a.address == local_ip:
                        return iface, local_ip
        return None, local_ip
    except: return None, "Unknown"

def get_gateway():
    try:
        if IS_WIN:
            out = subprocess.check_output("ipconfig", text=True,
                                          stderr=subprocess.DEVNULL)
            for line in out.split("\n"):
                if "Default Gateway" in line:
                    m = re.search(r"(\d{1,3}(?:\.\d{1,3}){3})", line)
                    if m: return m.group(1)
        elif IS_LIN:
            out = subprocess.check_output(
                ["ip", "route", "show", "default"],
                text=True, stderr=subprocess.DEVNULL)
            m = re.search(r"default via (\S+)", out)
            if m: return m.group(1)
        elif IS_MAC:
            out = subprocess.check_output(
                ["netstat", "-rn"], text=True, stderr=subprocess.DEVNULL)
            for line in out.split("\n"):
                if line.startswith("default"):
                    parts = line.split()
                    if len(parts) >= 2: return parts[1]
    except: pass
    return "Unknown"

def get_public_ip():
    try:
        import urllib.request
        with urllib.request.urlopen("https://api.ipify.org", timeout=5) as r:
            return r.read().decode().strip()
    except: return "Unavailable"

def get_local_subnet(ip):
    try:
        parts = ip.split(".")
        return f"{parts[0]}.{parts[1]}.{parts[2]}.0/24"
    except: return None

# ── Ping & Connectivity ───────────────────────────────────────────────────────
def ping_host(host, count=3):
    latencies, lost = [], 0
    for _ in range(count):
        cmd = (["ping", "-n", "1", "-w", "1000", host] if IS_WIN
               else ["ping", "-c", "1", "-W", "1", host])
        try:
            r = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                               text=True, timeout=3)
            if r.returncode == 0:
                m = re.search(r"time[=<]\s*(\d+\.?\d*)", r.stdout)
                if m: latencies.append(float(m.group(1)))
                else: lost += 1
            else: lost += 1
        except subprocess.TimeoutExpired:
            lost += 1
        except Exception:
            lost += 1
    avg    = sum(latencies) / len(latencies) if latencies else None
    loss   = (lost / count) * 100
    jitter = statistics.stdev(latencies) if len(latencies) >= 2 else 0.0
    return avg, loss, jitter

def check_dns():
    try:
        # Create a temporary socket with timeout instead of setting global default
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(3)
        s.close()
        # Use getaddrinfo which respects socket timeout better
        socket.getaddrinfo("google.com", 80, socket.AF_INET, socket.SOCK_STREAM)
        return True
    except Exception:
        return False

def get_network_status(gateway, targets, ping_count):
    res = {
        "status": "UNKNOWN", "gateway_ok": False, "internet_ok": False,
        "dns_ok": False, "latency": None, "loss": 100.0,
        "jitter": 0.0, "score": 0, "score_label": "Critical",
        "target_results": {}
    }
    if gateway and gateway != "Unknown":
        _, gw_loss, _ = ping_host(gateway, count=2)
        res["gateway_ok"] = gw_loss < 50

    all_lat, all_loss = [], []
    for t in targets:
        lat, loss, jitter = ping_host(t, count=ping_count)
        res["target_results"][t] = {"latency": lat, "loss": loss, "jitter": jitter}
        if lat is not None: all_lat.append(lat)
        all_loss.append(loss)

    if all_lat:
        res["latency"] = sum(all_lat) / len(all_lat)
        res["jitter"]  = statistics.stdev(all_lat) if len(all_lat) >= 2 else 0.0
    avg_loss = sum(all_loss) / len(all_loss) if all_loss else 100.0
    res["loss"]        = avg_loss
    res["internet_ok"] = avg_loss < 50
    res["dns_ok"]      = check_dns()

    if not res["gateway_ok"] and not res["internet_ok"]:
        res["status"] = "OFFLINE"
    elif res["gateway_ok"] and not res["internet_ok"]:
        res["status"] = "ISP_FAILURE"
    elif res["internet_ok"] and (avg_loss > 10 or
         (res["latency"] and res["latency"] > 200)):
        res["status"] = "DEGRADED"
    elif res["internet_ok"]:
        res["status"] = "ONLINE"
    else:
        res["status"] = "OFFLINE"

    res["score"]       = quality_score(res["latency"], res["loss"], res["jitter"])
    res["score_label"] = score_label(res["score"])
    return res

# ── Traffic Monitor ───────────────────────────────────────────────────────────
class TrafficMonitor:
    def __init__(self, iface=None):
        self.iface    = iface
        self._lock    = threading.Lock()
        self._last    = self._io()
        self._last_t  = time.time()
        self.up_speed = 0.0
        self.dn_speed = 0.0
        self.total_sent = 0
        self.total_recv = 0

    def _io(self):
        if not PSUTIL_OK: return None
        try:
            if self.iface:
                return psutil.net_io_counters(pernic=True).get(self.iface)
            return psutil.net_io_counters()
        except: return None

    def update(self):
        cur = self._io(); now = time.time()
        if cur is None or self._last is None: return
        dt = now - self._last_t
        if dt <= 0: return
        with self._lock:
            self.up_speed   = (cur.bytes_sent - self._last.bytes_sent) / dt
            self.dn_speed   = (cur.bytes_recv - self._last.bytes_recv) / dt
            self.total_sent = cur.bytes_sent
            self.total_recv = cur.bytes_recv
        self._last = cur; self._last_t = now

    def snap(self):
        with self._lock:
            return {"up": self.up_speed, "dn": self.dn_speed,
                    "sent": self.total_sent, "recv": self.total_recv}

# ── ARP Scanner ───────────────────────────────────────────────────────────────
class ARPScanner:
    OUI = {
        "00:50:56": "VMware",       "00:0C:29": "VMware",
        "DC:A6:32": "Raspberry Pi", "B8:27:EB": "Raspberry Pi",
        "52:54:00": "QEMU/KVM",     "00:25:00": "Apple",
        "AC:DE:48": "Apple",        "A4:C3:F0": "Apple",
        "00:50:F2": "Microsoft",    "28:D2:44": "Samsung",
        "00:26:B9": "Dell",         "18:03:73": "Dell",
        "00:21:70": "Cisco",        "00:1B:D4": "Cisco",
        "EC:08:6B": "TP-Link",      "50:C7:BF": "TP-Link",
        "14:CC:20": "TP-Link",      "00:17:88": "Philips Hue",
        "B4:E6:2D": "Netgear",      "C8:3A:35": "Tenda",
    }
    EXCLUDED = [
        ipaddress.ip_network("224.0.0.0/4"),
        ipaddress.ip_network("239.0.0.0/8"),
        ipaddress.ip_network("169.254.0.0/16"),
        ipaddress.ip_network("255.255.255.255/32"),
        ipaddress.ip_network("0.0.0.0/8"),
    ]

    def __init__(self, enabled=True, local_ip=None):
        self.enabled   = enabled
        self.local_ip  = local_ip
        self.subnet    = get_local_subnet(local_ip) if local_ip else None
        self.devices   = []
        self.last_scan = 0
        self._lock     = threading.Lock()
        self._running  = False

    def _valid_ip(self, ip):
        try:
            addr = ipaddress.ip_address(ip)
            with self._lock:
                local_ip = self.local_ip
                subnet = self.subnet
            if ip == local_ip: return False
            for net in self.EXCLUDED:
                if addr in net: return False
            if subnet:
                if addr not in ipaddress.ip_network(subnet, strict=False):
                    return False
            return True
        except: return False

    def _vendor(self, mac):
        return self.OUI.get(mac.upper()[:8], "Unknown")

    def _ping_check(self, ip):
        cmd = (["ping", "-n", "1", "-w", "500", ip] if IS_WIN
               else ["ping", "-c", "1", "-W", "1", ip])
        try:
            r = subprocess.run(cmd, stdout=subprocess.DEVNULL,
                               stderr=subprocess.DEVNULL, timeout=2)
            return r.returncode == 0
        except: return False

    def _scan(self):
        if not self.enabled: return []
        found, seen = [], set()
        try:
            out = subprocess.check_output(
                ["arp", "-a"], text=True,
                stderr=subprocess.DEVNULL, timeout=15)
            ip_re  = re.compile(r"\b(\d{1,3}(?:\.\d{1,3}){3})\b")
            mac_re = re.compile(
                r"\b([0-9a-fA-F]{2}[:\-][0-9a-fA-F]{2}[:\-][0-9a-fA-F]{2}"
                r"[:\-][0-9a-fA-F]{2}[:\-][0-9a-fA-F]{2}[:\-][0-9a-fA-F]{2})\b")
            for line in out.split("\n"):
                im = ip_re.search(line); mm = mac_re.search(line)
                if not im or not mm: continue
                ip  = im.group(1)
                mac = mm.group(1).upper().replace("-", ":")
                if not self._valid_ip(ip): continue
                if mac in ("FF:FF:FF:FF:FF:FF", "00:00:00:00:00:00"): continue
                if ip in seen: continue
                seen.add(ip)
                found.append({"ip": ip, "mac": mac,
                              "vendor": self._vendor(mac),
                              "alive": self._ping_check(ip)})
        except: pass
        return found

    def start(self, interval=60):
        self._running = True
        def loop():
            while self._running and not _shutdown_event.is_set():
                d = self._scan()
                with self._lock:
                    self.devices   = d
                    self.last_scan = time.time()
                _shutdown_event.wait(interval)
        threading.Thread(target=loop, daemon=True).start()

    def stop(self): self._running = False

    def get(self):
        with self._lock:
            return list(self.devices), self.last_scan

# ── Event Log (in-memory) ─────────────────────────────────────────────────────
class EventLog:
    def __init__(self, maxlen=500, max_age=3600):
        self._lock  = threading.Lock()
        self.events = []
        self.maxlen = maxlen
        self.max_age = max_age

    def log(self, event, detail="", prev="", new="", duration=None):
        rec = {
            "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "event":        event,
            "detail":       detail,
            "prev_status":  prev,
            "new_status":   new,
            "duration_sec": duration or ""
        }
        with self._lock:
            self.events.append(rec)
            now = time.time()
            cutoff = now - self.max_age
            self.events = [e for e in self.events if datetime.strptime(e["timestamp"], "%Y-%m-%d %H:%M:%S").timestamp() > cutoff]
            if len(self.events) > self.maxlen:
                self.events = self.events[-self.maxlen:]

    def recent(self, n=8):
        with self._lock: return list(self.events)[-n:]

    def all(self):
        with self._lock: return list(self.events)

# ── Metrics Store (in-memory) ─────────────────────────────────────────────────
class MetricsStore:
    def __init__(self, maxlen=120, max_age=3600):
        self._lock   = threading.Lock()
        self.records = []
        self.maxlen = maxlen
        self.max_age = max_age

    def add(self, net, traffic):
        rec = {
            "ts":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status":  net["status"],
            "latency": round(net["latency"], 2) if net["latency"] else None,
            "loss":    round(net["loss"], 2),
            "jitter":  round(net["jitter"], 2),
            "score":   net["score"],
            "up_bps":  round(traffic["up"], 2),
            "dn_bps":  round(traffic["dn"], 2),
            "dns_ok":  int(net["dns_ok"]),
            "gw_ok":   int(net["gateway_ok"]),
        }
        with self._lock:
            self.records.append(rec)
            now = time.time()
            cutoff = now - self.max_age
            self.records = [r for r in self.records if datetime.strptime(r["ts"], "%Y-%m-%d %H:%M:%S").timestamp() > cutoff]
            if len(self.records) > self.maxlen:
                self.records = self.records[-self.maxlen:]

    def all(self):
        with self._lock: return list(self.records)

    def all_locked(self):
        with self._lock:
            return list(self.records)

# ── Session Tracker ───────────────────────────────────────────────────────────
class Session:
    def __init__(self):
        self.start      = time.time()
        self.outages    = 0
        self.downtime   = 0.0
        self._out_start = None
        self._last_st   = None
        self._durations = []
        self._lock      = threading.Lock()

    def update(self, status, log: EventLog):
        now    = time.time()
        online = status in ("ONLINE", "DEGRADED")
        with self._lock:
            prev = self._last_st
            if prev is None:
                self._last_st = status
                log.log("SESSION_START", new=status, detail="Monitoring started")
                return
            if prev in ("ONLINE", "DEGRADED") and not online:
                self.outages    += 1
                self._out_start  = now
                self._last_st    = status
                log.log("OUTAGE_START", prev=prev, new=status,
                        detail=f"Network went {status}")
                beep()
            elif prev not in ("ONLINE", "DEGRADED") and online:
                dur = 0.0
                if self._out_start:
                    dur = round(now - self._out_start, 1)
                    self.downtime += dur
                    self._durations.append(dur)
                    self._out_start = None
                self._last_st = status
                log.log("RECOVERY", prev=prev, new=status,
                        detail="Network recovered", duration=dur)
                beep()
            elif prev != status:
                self._last_st = status
                log.log("STATUS_CHANGE", prev=prev, new=status,
                        detail=f"{prev} → {status}")

    def summary(self):
        with self._lock:
            now  = time.time()
            tot  = now - self.start
            down = self.downtime + (now - self._out_start if self._out_start else 0)
            up   = max(0.0, tot - down)
            pct  = (up / tot * 100) if tot > 0 else 100.0
            mtbo = statistics.mean(self._durations) if self._durations else None
            return {"total": tot, "uptime": up, "downtime": down,
                    "up_pct": pct, "outages": self.outages, "mtbo": mtbo}

# ── Pause Controller ──────────────────────────────────────────────────────────
class PauseController:
    def __init__(self, state):
        self._state = state
        self.paused   = False
        self._running = True
        self._lock    = threading.Lock()

    def start(self):
        threading.Thread(target=self._listen, daemon=True).start()

    def _listen(self):
        global BEEP_ENABLED
        if IS_WIN:
            import msvcrt
            while self._running and not _shutdown_event.is_set():
                if msvcrt.kbhit():
                    try:
                        ch = msvcrt.getch().decode("utf-8", "ignore").lower()
                        if ch == "p":
                            self.toggle()
                        elif ch == "b":
                            with _beep_lock:
                                BEEP_ENABLED = not BEEP_ENABLED
                                new_val = BEEP_ENABLED
                            with self._state["lock"]:
                                self._state["beep_enabled"] = new_val
                    except (UnicodeDecodeError, AttributeError):
                        pass  # Ignore special keys that can't be decoded
                _shutdown_event.wait(0.05)
        else:
            import tty, termios, select
            fd  = sys.stdin.fileno()
            old = termios.tcgetattr(fd)
            def restore_terminal():
                try:
                    termios.tcsetattr(fd, termios.TCSADRAIN, old)
                except Exception:
                    pass
            atexit.register(restore_terminal)
            try:
                tty.setraw(fd)
                while self._running and not _shutdown_event.is_set():
                    if select.select([sys.stdin], [], [], 0.05)[0]:
                        ch = sys.stdin.read(1).lower()
                        if ch == "p":
                            self.toggle()
                        elif ch == "b":
                            with _beep_lock:
                                BEEP_ENABLED = not BEEP_ENABLED
                                new_val = BEEP_ENABLED
                            with self._state["lock"]:
                                self._state["beep_enabled"] = new_val
            finally:
                restore_terminal()

    def is_paused(self):
        with self._lock: return self.paused

    def toggle(self):
        with self._lock:
            self.paused = not self.paused
            return self.paused

    def stop(self): self._running = False

# ── Terminal Dashboard ────────────────────────────────────────────────────────
class TerminalDashboard:
    W = 62

    def _hdr(self, title):
        pad = self.W - len(title) - 4
        return cl(f"╭─ {title} " + "─" * pad + "╮", A.CYN, A.BOLD)

    def _ftr(self):
        return cl("╰" + "─" * (self.W - 2) + "╯", A.CYN)

    def _row(self, label, val, vc=None):
        lbl = cl(f"  {label:<22}", A.GRY)
        v   = cl(str(val), vc) if vc else str(val)
        return f"{lbl}{v}"

    def draw(self, net, traffic, devices, last_scan,
             local_ip, public_ip, gateway, iface,
             session, log, metrics_store, lat_hist,
             paused, no_priv, args, beep_enabled, live_interval):

        lines = []
        add   = lines.append

        status = net["status"]
        lat    = net["latency"]
        loss   = net["loss"]
        jit    = net["jitter"]
        score  = net["score"]
        sc     = status_col(status)
        now    = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

        # ── Title ────────────────────────────────────────────────────────────
        add("")
        add(cl("  ◈ ZYTRONET", A.CYN, A.BOLD) +
            cl(" v3.0", A.GRY) + "  " + cl(now, A.GRY))
        add(cl("  " + "═" * self.W, A.CYN))
        if paused:
            add(cl("  ⏸  PAUSED — press P to resume", A.YLW, A.BOLD))
        add("")

        # ── Network Status ────────────────────────────────────────────────────
        add(self._hdr("Network Status"))
        add(cl(f"  ● {status}", sc, A.BOLD))

        lat_c = A.GRN if lat and lat < 80 else A.YLW if lat and lat < 200 else A.RED
        add(self._row("Latency",     f"{lat:.1f} ms" if lat else "N/A", lat_c))
        add(self._row("Packet Loss", f"{loss:.1f}%",
            A.GRN if loss == 0 else A.YLW if loss < 15 else A.RED))
        add(self._row("Jitter",      f"{jit:.1f} ms",
            A.GRN if jit < 10 else A.YLW if jit < 30 else A.RED))
        add(self._row("DNS",
            cl("Resolving ✓", A.GRN) if net["dns_ok"] else cl("Failed ✗", A.RED)))

        if not net["gateway_ok"] and not net["internet_ok"]:
            fault = cl("Local network failure", A.RED)
        elif net["gateway_ok"] and not net["internet_ok"]:
            fault = cl("ISP / upstream failure", A.YLW)
        else:
            fault = cl("None detected", A.GRN)
        add(self._row("Fault Location", fault))

        sc_term = A.GRN if score >= 85 else A.YLW if score >= 50 else A.RED
        add(self._row("Quality Score",
            cl(f"{score}/100  {net['score_label']}", sc_term, A.BOLD)))
        add(self._ftr())
        add("")

        # ── Latency Sparkline ─────────────────────────────────────────────────
        if lat_hist:
            vals = list(lat_hist)
            mn, mx, avg = min(vals), max(vals), sum(vals) / len(vals)
            add(cl("  Latency History", A.GRY, A.BOLD))
            add(f"  {sparkline(vals)}")
            add(f"  {cl(f'min {mn:.0f}ms', A.GRN)}   "
                f"{cl(f'avg {avg:.0f}ms', A.CYN)}   "
                f"{cl(f'max {mx:.0f}ms', A.RED)}")
            add("")

        # ── Target Results ────────────────────────────────────────────────────
        add(self._hdr("Target Results"))
        for t, r in net["target_results"].items():
            tl  = f"{r['latency']:.1f} ms" if r["latency"] else "Timeout"
            tlc = A.GRN if r["latency"] and r["latency"] < 100 else A.RED
            add(self._row(t,
                cl(tl, tlc) +
                cl(f"  jitter {r['jitter']:.1f}ms  loss {r['loss']:.0f}%", A.GRY)))
        add(self._ftr())
        add("")

        # ── Traffic ───────────────────────────────────────────────────────────
        add(self._hdr(f"Traffic  [{iface or 'Unknown'}]"))
        add(self._row("↑ Upload Speed",   fmt_speed(traffic["up"]),   A.YLW))
        add(self._row("↓ Download Speed", fmt_speed(traffic["dn"]),   A.GRN))
        add(self._row("Total Sent",       fmt_bytes(traffic["sent"]), A.GRY))
        add(self._row("Total Received",   fmt_bytes(traffic["recv"]), A.GRY))
        add(self._ftr())
        add("")

        # ── System ────────────────────────────────────────────────────────────
        add(self._hdr("System"))
        add(self._row("Local IP",   local_ip,  A.CYN))
        add(self._row("Public IP",  public_ip, A.MGT))
        add(self._row("Gateway",    gateway,   A.CYN))
        add(self._ftr())
        add("")

        # ── Devices ───────────────────────────────────────────────────────────
        add(self._hdr("Connected Devices"))
        if args.no_arp:
            add(cl("  ARP scanning disabled", A.GRY))
        elif no_priv:
            add(cl("  ⚠  Run as admin for full device discovery", A.YLW))
        if devices:
            age = int(time.time() - last_scan) if last_scan else 0
            add(cl(f"  {'IP':<18}{'Vendor':<16}Status", A.GRY))
            add(cl("  " + "─" * (self.W - 4), A.GRY))
            for d in devices[:6]:
                alive = cl("● Online", A.GRN) if d["alive"] else cl("○ Stale", A.GRY)
                add(f"  {cl(d['ip'], A.WHT):<28}"
                    f"{cl(d['vendor'], A.GRY):<16}{alive}")
            add(cl(f"\n  {len(devices)} device(s)", A.WHT) +
                cl(f"   scan {age}s ago", A.GRY))
        else:
            add(cl("  Scanning...", A.GRY))
        add(self._ftr())
        add("")

        # ── Session ───────────────────────────────────────────────────────────
        ss = session.summary()
        add(self._hdr("Session"))
        add(self._row("Runtime",  fmt_dur(ss["total"]),    A.WHT))
        add(self._row("Uptime",   fmt_dur(ss["uptime"]),   A.GRN))
        add(self._row("Downtime", fmt_dur(ss["downtime"]),
            A.RED if ss["downtime"] > 0 else A.GRY))
        up_c = A.GRN if ss["up_pct"] > 95 else A.YLW if ss["up_pct"] > 80 else A.RED
        add(self._row("Uptime %", f"{ss['up_pct']:.2f}%", up_c))
        add(self._row("Outages",  str(ss["outages"]),
            A.RED if ss["outages"] > 0 else A.GRN))
        if ss["mtbo"]:
            add(self._row("Avg Outage", fmt_dur(ss["mtbo"]), A.YLW))
        add(self._ftr())
        add("")

        # ── Recent Events ─────────────────────────────────────────────────────
        add(self._hdr("Recent Events"))
        evts = log.recent(4)
        if evts:
            for ev in reversed(evts):
                add(f"  {cl(ev['timestamp'], A.GRY)}  "
                    f"{cl(ev['event'][:16], A.CYN):<24}  "
                    f"{cl(ev['detail'], A.WHT)}")
        else:
            add(cl("  No events yet.", A.GRY))
        add(self._ftr())
        add("")

        # ── Footer ────────────────────────────────────────────────────────────
        mc = len(metrics_store.all())
        add(cl(f"  {mc} samples", A.GRY) +
            cl("  │  P = pause/resume", A.GRY) +
            cl("  │  " + ("[B = mute beep]" if beep_enabled else "[B = unmute beep] 🔇"), A.GRY) +
            cl(f"  │  Interval: {live_interval:.1f}s", A.GRY) +
            cl("  │  CTRL+C = stop + report", A.GRY))
        if not args.no_web and FLASK_OK:
            add(cl(f"  Live web dashboard → http://localhost:{args.web_port}", A.CYN))
        add("")

        # ── Render (clear + redraw) ───────────────────────────────────────────
        os.system('cls' if os.name == 'nt' else 'clear')
        print("\n".join(lines))


def start_interface_watcher(state, event_log, scanner):
    def loop():
        while not _shutdown_event.is_set():
            new_iface, new_local_ip = get_primary_interface()
            with state["lock"]:
                current_local_ip = state.get("local_ip", "Unknown")
            if new_local_ip == "Unknown" or new_local_ip == current_local_ip:
                _shutdown_event.wait(5)
                continue

            with state["lock"]:
                old_local_ip = state.get("local_ip", "Unknown")
                state["public_ip"] = "Fetching..."

            new_public_ip = get_public_ip()
            new_gateway = get_gateway()
            new_subnet = get_local_subnet(new_local_ip)

            with state["lock"]:
                state["local_ip"] = new_local_ip
                state["public_ip"] = new_public_ip
                state["gateway"] = new_gateway
                state["iface"] = new_iface

            with scanner._lock:
                scanner.local_ip = new_local_ip
                scanner.subnet = new_subnet

            event_log.log(
                "NETWORK_CHANGE",
                detail=f"{old_local_ip} -> {new_local_ip}",
                prev=old_local_ip,
                new=new_local_ip,
            )
            _shutdown_event.wait(5)

    threading.Thread(target=loop, daemon=True).start()


# ── Web Dashboard HTML ────────────────────────────────────────────────────────
WEB_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>ZytroNet Live</title>
<meta name="author" content="Elkmirl Kuuku Carter">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root{
  color-scheme:dark;
  --bg:#07111f;
  --bg-elev:#0b1628;
  --panel:#0f1b31cc;
  --panel-strong:#12213d;
  --panel-soft:#162742;
  --border:#22314b;
  --border-strong:#324562;
  --text:#f4f7fb;
  --text-soft:#b6c3d9;
  --text-faint:#7f92b1;
  --accent:#79a8ff;
  --accent-strong:#4f7fff;
  --accent-soft:rgba(121,168,255,.16);
  --success:#33d69f;
  --success-soft:rgba(51,214,159,.16);
  --warning:#f6c760;
  --warning-soft:rgba(246,199,96,.16);
  --danger:#ff6b81;
  --danger-soft:rgba(255,107,129,.16);
  --purple:#b388ff;
  --shadow-lg:0 24px 80px rgba(0,0,0,.35);
  --shadow-md:0 16px 40px rgba(2,8,23,.28);
  --radius-xl:24px;
  --radius-lg:18px;
  --radius-md:14px;
  --radius-sm:10px;
  --space-1:4px;
  --space-2:8px;
  --space-3:12px;
  --space-4:16px;
  --space-5:20px;
  --space-6:24px;
  --space-8:32px;
  --space-10:40px;
}
body[data-theme="light"]{
  color-scheme:light;
  --bg:#f4f7fb;
  --bg-elev:#ffffff;
  --panel:rgba(255,255,255,.88);
  --panel-strong:#ffffff;
  --panel-soft:#edf3ff;
  --border:#dce4f1;
  --border-strong:#c9d5e8;
  --text:#0f172a;
  --text-soft:#475569;
  --text-faint:#64748b;
  --accent:#2563eb;
  --accent-strong:#1d4ed8;
  --accent-soft:rgba(37,99,235,.1);
  --success:#0f9f6e;
  --success-soft:rgba(15,159,110,.1);
  --warning:#c58a16;
  --warning-soft:rgba(197,138,22,.12);
  --danger:#d63e57;
  --danger-soft:rgba(214,62,87,.12);
  --purple:#7c3aed;
  --shadow-lg:0 24px 60px rgba(15,23,42,.12);
  --shadow-md:0 12px 30px rgba(15,23,42,.08);
}
*{box-sizing:border-box}
html,body{min-height:100%}
body{
  margin:0;
  color:var(--text);
  font-family:Inter,Segoe UI,Arial,sans-serif;
  background:
    radial-gradient(circle at top left, rgba(121,168,255,.16), transparent 24%),
    radial-gradient(circle at top right, rgba(179,136,255,.12), transparent 20%),
    linear-gradient(180deg, var(--bg) 0%, var(--bg-elev) 100%);
  transition:background .35s ease,color .25s ease;
}
a{color:inherit;text-decoration:none}
button,input{font:inherit}
#status-bg{
  position:fixed;inset:0;pointer-events:none;z-index:0;opacity:1;
  transition:background .45s ease, opacity .45s ease;
}
.shell,.alert-banner,.paused-banner,.modal-overlay{position:relative;z-index:1}
.shell{max-width:1480px;margin:0 auto;padding:24px 24px 32px}
.topbar{
  display:flex;justify-content:space-between;align-items:center;gap:16px;
  padding:18px 22px;margin-bottom:20px;border:1px solid var(--border);
  background:var(--panel);backdrop-filter:blur(16px);border-radius:22px;box-shadow:var(--shadow-md);
}
.brand{display:flex;align-items:center;gap:14px;min-width:0}
.brand-mark{
  width:42px;height:42px;border-radius:14px;display:grid;place-items:center;
  background:linear-gradient(135deg,var(--accent),var(--purple));color:#fff;
  font-size:18px;font-weight:800;box-shadow:0 12px 28px rgba(79,127,255,.32);
}
.brand-copy{min-width:0}
.eyebrow{
  display:inline-flex;align-items:center;gap:8px;padding:6px 10px;
  border-radius:999px;background:var(--accent-soft);color:var(--accent);
  font-size:12px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;
}
.eyebrow .live-dot{
  width:7px;height:7px;border-radius:999px;background:currentColor;
  box-shadow:0 0 0 6px rgba(121,168,255,.12);
}
.brand-copy h1{margin:10px 0 4px;font-size:20px;line-height:1.1}
.brand-copy p{margin:0;color:var(--text-soft);font-size:13px}
.toolbar{display:flex;flex-wrap:wrap;justify-content:flex-end;gap:10px}
.btn,.chip-btn,.event-filter-btn{
  display:inline-flex;align-items:center;justify-content:center;gap:8px;
  border:1px solid var(--border);border-radius:12px;padding:10px 14px;
  background:var(--panel-strong);color:var(--text);cursor:pointer;
  transition:transform .18s ease, border-color .18s ease, background .18s ease, color .18s ease, box-shadow .18s ease;
  box-shadow:none;
}
.btn:hover,.chip-btn:hover,.event-filter-btn:hover{
  transform:translateY(-1px);border-color:var(--border-strong);box-shadow:var(--shadow-md);
}
.btn:active,.chip-btn:active,.event-filter-btn:active{transform:translateY(0)}
.btn svg,.chip-btn svg{width:16px;height:16px}
.btn-primary{
  border-color:transparent;background:linear-gradient(135deg,var(--accent),var(--accent-strong));color:#fff;
}
.btn-tonal{background:var(--accent-soft);color:var(--accent);border-color:transparent}
.btn-ghost{background:transparent;color:var(--text-soft)}
.btn-alert{background:var(--danger-soft);color:var(--danger);border-color:transparent}
.paused-banner,.alert-banner{
  display:none;align-items:center;justify-content:space-between;gap:16px;
  margin-bottom:16px;padding:14px 18px;border-radius:16px;border:1px solid var(--border);
  backdrop-filter:blur(14px);box-shadow:var(--shadow-md);
}
.paused-banner{background:var(--warning-soft);border-color:rgba(246,199,96,.35);color:var(--warning)}
.alert-banner{background:var(--danger-soft);border-color:rgba(255,107,129,.35);color:var(--danger)}
.hero{
  display:grid;grid-template-columns:minmax(0,1.45fr) minmax(320px,.95fr);gap:18px;margin-bottom:18px;
}
.hero-card,.side-panel,.panel,.chart-card,.table-card,.map-card{
  background:var(--panel);border:1px solid var(--border);border-radius:var(--radius-xl);
  box-shadow:var(--shadow-md);backdrop-filter:blur(16px);
}
.hero-card{padding:24px}
.hero-top{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;margin-bottom:18px}
.hero-copy h2{margin:0 0 8px;font-size:28px;line-height:1.08;letter-spacing:-.03em}
.hero-copy p{margin:0;color:var(--text-soft);max-width:60ch}
.status-pill{
  display:inline-flex;align-items:center;gap:8px;padding:9px 14px;border-radius:999px;
  background:var(--panel-soft);border:1px solid var(--border);font-weight:700;
}
.status-pill::before{
  content:"";width:9px;height:9px;border-radius:999px;background:var(--text-faint);
  box-shadow:0 0 0 5px rgba(148,163,184,.12);
}
.status-pill[data-status="ONLINE"]{color:var(--success)}
.status-pill[data-status="ONLINE"]::before{background:var(--success);box-shadow:0 0 0 5px rgba(51,214,159,.18)}
.status-pill[data-status="DEGRADED"]{color:var(--warning)}
.status-pill[data-status="DEGRADED"]::before{background:var(--warning);box-shadow:0 0 0 5px rgba(246,199,96,.18)}
.status-pill[data-status="ISP_FAILURE"],.status-pill[data-status="OFFLINE"]{color:var(--danger)}
.status-pill[data-status="ISP_FAILURE"]::before,.status-pill[data-status="OFFLINE"]::before{background:var(--danger);box-shadow:0 0 0 5px rgba(255,107,129,.18)}
.stats-grid{
  display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;
}
.stat{
  padding:16px;border:1px solid var(--border);border-radius:18px;background:linear-gradient(180deg,var(--panel-soft),transparent);
}
.stat-label{font-size:12px;color:var(--text-faint);text-transform:uppercase;letter-spacing:.05em;font-weight:700}
.stat-value{margin-top:12px;font-size:28px;line-height:1;font-weight:800;letter-spacing:-.03em}
.stat-sub{margin-top:8px;color:var(--text-soft);font-size:13px}
.side-panel{padding:22px;display:flex;flex-direction:column;gap:18px}
.section-title{
  display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:12px;
}
.section-title h3,.panel h3,.chart-card h3,.table-card h3,.map-card h3{
  margin:0;font-size:16px;letter-spacing:-.02em;
}
.muted{color:var(--text-soft)}
.small{font-size:13px}
.tiny{font-size:12px;color:var(--text-faint)}
.stack{display:flex;flex-direction:column;gap:14px}
.kv{
  display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px;
}
.kv-item{
  padding:14px;border-radius:16px;background:var(--panel-soft);border:1px solid var(--border);
}
.kv-label{font-size:11px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--text-faint)}
.kv-value{margin-top:8px;font-size:14px;font-weight:700;word-break:break-word}
.control-card{
  padding:16px;border-radius:18px;background:var(--panel-soft);border:1px solid var(--border);display:flex;flex-direction:column;gap:12px;
}
.range-row,.input-row,.helper-row{display:flex;gap:10px;align-items:center}
.input,.range{
  width:100%;border-radius:12px;border:1px solid var(--border);background:var(--panel-strong);color:var(--text);
}
.input{padding:11px 13px;outline:none;transition:border-color .18s ease, box-shadow .18s ease}
.input:focus{border-color:var(--accent);box-shadow:0 0 0 4px rgba(121,168,255,.14)}
.range{appearance:none;height:8px;padding:0;background:linear-gradient(90deg,var(--accent),var(--purple))}
.range::-webkit-slider-thumb{
  appearance:none;width:18px;height:18px;border-radius:999px;border:0;background:#fff;box-shadow:0 6px 18px rgba(15,23,42,.3);cursor:pointer;
}
.range::-moz-range-thumb{
  width:18px;height:18px;border:0;border-radius:999px;background:#fff;box-shadow:0 6px 18px rgba(15,23,42,.3);cursor:pointer;
}
.error{min-height:18px;color:var(--danger);font-size:12px}
.targets-list{display:flex;flex-direction:column;gap:10px}
.target-item{
  display:flex;justify-content:space-between;align-items:center;gap:12px;
  padding:12px 14px;border-radius:14px;background:var(--panel-strong);border:1px solid var(--border);
}
.target-ip{font-weight:700}
.dashboard-grid{
  display:grid;grid-template-columns:minmax(0,1.15fr) minmax(0,1.15fr) minmax(320px,.8fr);gap:18px;
}
.charts-column,.events-column{display:flex;flex-direction:column;gap:18px}
.chart-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:18px}
.chart-card,.table-card,.map-card{padding:20px}
.chart-meta,.table-meta,.map-meta{
  color:var(--text-faint);font-size:12px;margin-top:4px;
}
.chart-head,.table-head,.map-head{
  display:flex;justify-content:space-between;gap:12px;align-items:flex-start;margin-bottom:14px;
}
.card-badge{
  display:inline-flex;align-items:center;gap:6px;padding:7px 10px;border-radius:999px;
  font-size:12px;font-weight:700;border:1px solid var(--border);background:var(--panel-soft);color:var(--text-soft);
}
.chart-wrap{height:240px}
.chart-wrap canvas{width:100% !important;height:100% !important;cursor:crosshair}
.map-wrap{
  padding:12px;border-radius:18px;background:linear-gradient(180deg,var(--panel-soft),transparent);
  border:1px solid var(--border);
}
#net-map{width:100%;height:auto;display:block}
.filter-row{
  display:flex;flex-wrap:wrap;gap:8px;margin-bottom:14px;
}
.event-filter-btn{
  background:var(--panel-soft);color:var(--text-soft);padding:8px 12px;border-radius:999px;
}
.event-filter-btn.active-filter{background:var(--accent-soft);border-color:transparent;color:var(--accent)}
.notif-banner{
  display:none;align-items:center;justify-content:space-between;gap:12px;
  margin-bottom:14px;padding:12px 14px;border-radius:14px;background:var(--panel-soft);
  border:1px solid var(--border);color:var(--text-soft);cursor:pointer;
}
.table-shell{
  overflow:auto;border-radius:18px;border:1px solid var(--border);background:var(--panel-soft);
}
table{width:100%;border-collapse:collapse;min-width:640px}
thead th{
  position:sticky;top:0;background:var(--panel-strong);z-index:1;
  color:var(--text-faint);font-size:11px;font-weight:800;letter-spacing:.06em;text-transform:uppercase;
}
th,td{padding:14px 16px;border-bottom:1px solid var(--border);text-align:left;font-size:13px;vertical-align:top}
tbody tr{transition:background .18s ease}
tbody tr:hover{background:rgba(121,168,255,.06)}
.badge{
  display:inline-flex;align-items:center;padding:6px 10px;border-radius:999px;font-size:11px;font-weight:800;letter-spacing:.04em;text-transform:uppercase;
}
.badge-online{background:var(--success-soft);color:var(--success)}
.badge-warn{background:var(--warning-soft);color:var(--warning)}
.badge-danger{background:var(--danger-soft);color:var(--danger)}
.badge-neutral{background:var(--panel-strong);color:var(--text-soft)}
.empty-state{
  display:flex;flex-direction:column;align-items:flex-start;gap:10px;
  padding:24px;border-radius:18px;border:1px dashed var(--border-strong);background:var(--panel-soft);color:var(--text-soft);
}
.skeleton{
  position:relative;overflow:hidden;border-radius:14px;background:var(--panel-soft);min-height:56px;
}
.skeleton::after{
  content:"";position:absolute;inset:0;transform:translateX(-100%);
  background:linear-gradient(90deg, transparent, rgba(255,255,255,.12), transparent);
  animation:loading 1.5s infinite;
}
@keyframes loading{100%{transform:translateX(100%)}}
.modal-overlay{
  display:none;position:fixed;inset:0;padding:24px;background:rgba(3,8,20,.55);
  backdrop-filter:blur(10px);align-items:center;justify-content:center;
}
.modal-box{
  width:min(980px,96vw);padding:22px;border-radius:24px;background:var(--panel-strong);
  border:1px solid var(--border);box-shadow:var(--shadow-lg);
}
.modal-box .chart-wrap{height:360px}
.footer{
  margin-top:18px;padding:16px 8px;color:var(--text-faint);font-size:12px;text-align:center;
}
@media (max-width:1280px){
  .dashboard-grid{grid-template-columns:1fr}
  .events-column{order:3}
}
@media (max-width:1040px){
  .hero{grid-template-columns:1fr}
  .stats-grid,.chart-grid,.kv{grid-template-columns:1fr 1fr}
}
@media (max-width:720px){
  .shell{padding:16px}
  .topbar,.hero-card,.side-panel,.chart-card,.table-card,.map-card{border-radius:20px}
  .topbar,.hero-top,.toolbar,.range-row,.input-row{flex-direction:column;align-items:stretch}
  .stats-grid,.chart-grid,.kv{grid-template-columns:1fr}
  .brand{align-items:flex-start}
  .btn,.chip-btn{width:100%}
}
</style>
</head>
<body data-theme="dark">
<div id="status-bg"></div>

<div class="shell">
  <div id="outage-alert" class="alert-banner">
    <div>
      <div class="eyebrow" style="margin-bottom:8px;background:var(--danger-soft);color:var(--danger)"><span class="live-dot"></span>Attention needed</div>
      <strong id="alert-text">Network outage detected</strong>
    </div>
    <button class="btn btn-alert" onclick="dismissAlert()">Dismiss</button>
  </div>

  <div id="paused-banner" class="paused-banner">
    <div>
      <strong>Monitoring paused</strong>
      <div class="small">Live polling is temporarily stopped. Resume to continue refreshing metrics and events.</div>
    </div>
    <button id="pause-banner-btn" class="btn btn-tonal" type="button">Resume monitoring</button>
  </div>

  <header class="topbar">
    <div class="brand">
      <div class="brand-mark">Z</div>
      <div class="brand-copy">
        <div class="eyebrow"><span class="live-dot"></span>Live monitoring</div>
        <h1>ZytroNet Network Operations</h1>
        <p id="upd">Connecting to monitoring engine...</p>
      </div>
    </div>
    <div class="toolbar">
      <a href="/report" target="_blank" download="zytronet_report.html" class="btn btn-primary">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M12 3v12m0 0 4-4m-4 4-4-4"/><path d="M4 17v1a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2v-1"/></svg>
        Export report
      </a>
      <button id="beep-btn" class="btn btn-ghost" type="button">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M15 8a5 5 0 0 1 0 8"/><path d="M17.7 5a9 9 0 0 1 0 14"/><path d="M5 10h3l4-4v12l-4-4H5z"/></svg>
        Alerts sound on
      </button>
      <button id="pause-btn" class="btn btn-ghost" type="button">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M8 5v14M16 5v14"/></svg>
        Pause monitoring
      </button>
      <button id="theme-btn" class="btn btn-ghost" type="button">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M12 3v2m0 14v2M5.64 5.64l1.41 1.41m9.9 9.9 1.41 1.41M3 12h2m14 0h2M5.64 18.36l1.41-1.41m9.9-9.9 1.41-1.41"/><circle cx="12" cy="12" r="4"/></svg>
        Toggle theme
      </button>
    </div>
  </header>

  <section class="hero">
    <div class="hero-card">
      <div class="hero-top">
        <div class="hero-copy">
          <h2>Network health, traffic, and incident visibility in one premium dashboard.</h2>
          <p>Track uptime, quality, latency, and discovered devices with live updates, richer hierarchy, and clearer operational feedback.</p>
        </div>
        <div id="status-pill" class="status-pill" data-status="UNKNOWN">Status unknown</div>
      </div>
      <div class="stats-grid">
        <div class="stat">
          <div class="stat-label">Latency</div>
          <div id="hlat" class="stat-value">--</div>
          <div class="stat-sub">Current round-trip response</div>
        </div>
        <div class="stat">
          <div class="stat-label">Packet Loss</div>
          <div id="hloss" class="stat-value">--</div>
          <div class="stat-sub">Latest monitoring window</div>
        </div>
        <div class="stat">
          <div class="stat-label">Uptime</div>
          <div id="hup" class="stat-value">--</div>
          <div class="stat-sub">Session availability</div>
        </div>
        <div class="stat">
          <div class="stat-label">Quality Score</div>
          <div id="hscore" class="stat-value">--</div>
          <div id="hscore-label" class="stat-sub">Waiting for metrics</div>
        </div>
      </div>
    </div>

    <aside class="side-panel">
      <div>
        <div class="section-title">
          <h3>Environment</h3>
          <span class="card-badge">Live context</span>
        </div>
        <div class="kv">
          <div class="kv-item"><div class="kv-label">Local IP</div><div id="lip" class="kv-value">--</div></div>
          <div class="kv-item"><div class="kv-label">Public IP</div><div id="pip" class="kv-value">--</div></div>
          <div class="kv-item"><div class="kv-label">Gateway</div><div id="gw" class="kv-value">--</div></div>
          <div class="kv-item"><div class="kv-label">Interface</div><div id="ifc" class="kv-value">--</div></div>
        </div>
      </div>

      <div class="control-card">
        <div class="section-title">
          <h3>Polling cadence</h3>
          <span id="interval-label" class="card-badge">--</span>
        </div>
        <div class="range-row">
          <input id="interval-slider" class="range" type="range" min="0.5" max="60" step="0.5">
        </div>
        <div class="helper-row tiny">Adjust refresh cadence without interrupting the active session.</div>
      </div>

      <div class="control-card">
        <div class="section-title">
          <h3>Targets</h3>
          <span id="target-count" class="card-badge">0 active</span>
        </div>
        <div class="input-row">
          <input id="target-input" class="input" placeholder="Add target, e.g. 8.8.8.8">
          <button class="btn btn-tonal" onclick="addTarget()" type="button">Add</button>
        </div>
        <div id="target-error" class="error"></div>
        <div id="target-list" class="targets-list">
          <div class="skeleton"></div>
        </div>
      </div>
    </aside>
  </section>

  <section class="dashboard-grid">
    <div class="charts-column" style="grid-column:span 2">
      <div class="chart-grid">
        <article class="chart-card">
          <div class="chart-head">
            <div>
              <h3>Latency trend</h3>
              <div class="chart-meta">Operational responsiveness across the latest samples.</div>
            </div>
            <span class="card-badge">Interactive</span>
          </div>
          <div class="chart-wrap"><canvas id="cLat"></canvas></div>
        </article>

        <article class="chart-card">
          <div class="chart-head">
            <div>
              <h3>Packet loss</h3>
              <div class="chart-meta">Visibility into instability and drop events.</div>
            </div>
            <span class="card-badge">Interactive</span>
          </div>
          <div class="chart-wrap"><canvas id="cLoss"></canvas></div>
        </article>

        <article class="chart-card">
          <div class="chart-head">
            <div>
              <h3>Traffic throughput</h3>
              <div class="chart-meta">Upstream and downstream movement in KB/s.</div>
            </div>
            <span class="card-badge">Interactive</span>
          </div>
          <div class="chart-wrap"><canvas id="cSpeed"></canvas></div>
        </article>

        <article class="chart-card">
          <div class="chart-head">
            <div>
              <h3>Quality score</h3>
              <div class="chart-meta">Composite signal based on latency, jitter, and loss.</div>
            </div>
            <span class="card-badge">Interactive</span>
          </div>
          <div class="chart-wrap"><canvas id="cScore"></canvas></div>
        </article>
      </div>

      <article class="map-card">
        <div class="map-head">
          <div>
            <h3>Network map</h3>
            <div class="map-meta">Discovered devices around the current gateway. Click any node to run an on-demand ping.</div>
          </div>
          <span id="device-count" class="card-badge">0 devices</span>
        </div>
        <div class="map-wrap">
          <canvas id="net-map" width="900" height="360"></canvas>
        </div>
      </article>
    </div>

    <div class="events-column">
      <article class="table-card">
        <div class="table-head">
          <div>
            <h3>Events</h3>
            <div class="table-meta">Incident history, recoveries, and session milestones.</div>
          </div>
          <span id="event-count" class="card-badge">0 entries</span>
        </div>
        <div id="event-filters" class="filter-row">
          <button class="event-filter-btn active-filter" data-filter="All" onclick="setEventFilter(this)" type="button">All</button>
          <button class="event-filter-btn" data-filter="OUTAGE_START" onclick="setEventFilter(this)" type="button">Outages</button>
          <button class="event-filter-btn" data-filter="RECOVERY" onclick="setEventFilter(this)" type="button">Recoveries</button>
          <button class="event-filter-btn" data-filter="STATUS_CHANGE" onclick="setEventFilter(this)" type="button">Status changes</button>
          <button class="event-filter-btn" data-filter="SESSION_START" onclick="setEventFilter(this)" type="button">Session start</button>
        </div>
        <div id="notif-banner" class="notif-banner">
          <div>
            <strong>Enable notifications</strong>
            <div class="tiny">Get browser alerts for outages and recoveries.</div>
          </div>
          <span class="card-badge">Optional</span>
        </div>
        <div class="table-shell">
          <table>
            <thead>
              <tr><th>Timestamp</th><th>Type</th><th>Detail</th><th>Duration</th></tr>
            </thead>
            <tbody id="evrows">
              <tr><td colspan="4"><div class="empty-state"><strong>Waiting for telemetry</strong><span>Events will appear here as the monitoring engine streams new activity.</span></div></td></tr>
            </tbody>
          </table>
        </div>
      </article>
    </div>
  </section>

  <footer class="footer">ZytroNet live dashboard · Developed by Elkmirl Kuuku Carter</footer>
</div>

<div id="modal-overlay" class="modal-overlay" onclick="closeModal()">
  <div class="modal-box" onclick="event.stopPropagation()">
    <div class="section-title" style="margin-bottom:18px">
      <div>
        <h3 id="modal-title">Detailed chart</h3>
        <div class="chart-meta">Expanded chart view for closer analysis.</div>
      </div>
      <button class="btn btn-ghost" onclick="closeModal()" type="button">Close</button>
    </div>
    <div class="chart-wrap"><canvas id="modal-canvas"></canvas></div>
  </div>
</div>

<script>
const SC={ONLINE:'#33d69f',DEGRADED:'#f6c760',ISP_FAILURE:'#ff8a65',OFFLINE:'#ff6b81',UNKNOWN:'#94a3b8'};
const SCORE_LABELS=[{min:85,label:'Excellent'},{min:70,label:'Good'},{min:50,label:'Fair'},{min:25,label:'Poor'},{min:0,label:'Critical'}];
function el(id){return document.getElementById(id)}
function escapeHtml(v){return String(v??'').replace(/[&<>"']/g,(c)=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]))}
function scoreLabel(score){const n=Number(score||0);for(const item of SCORE_LABELS){if(n>=item.min)return item.label}return 'Unknown'}
function eventBadgeCls(name){if(name==='RECOVERY')return 'badge badge-online';if(name==='OUTAGE_START'||name==='ISP_FAILURE')return 'badge badge-danger';if(name==='STATUS_CHANGE')return 'badge badge-warn';return 'badge badge-neutral'}
let latestMetrics=[],latestNet={},latestEvents=[],latestDevices=[],lastTargetsJson="",activeFilter="All";
let sliderDragging=false,lastGoodInterval=null,modalChart=null,lastEventTs=null,alertDismissed=false;
let nodeHitTargets=[],pingResults={},pendingPings=new Set();
function css(name){return getComputedStyle(document.body).getPropertyValue(name).trim()}
function chartOptions(yLabel){
  return {
    responsive:true,
    maintainAspectRatio:false,
    animation:false,
    interaction:{mode:'index',intersect:false},
    plugins:{
      legend:{display:true,labels:{color:css('--text-soft'),boxWidth:10,boxHeight:10,usePointStyle:true,pointStyle:'circle'}},
      tooltip:{
        enabled:true,
        mode:'index',
        intersect:false,
        backgroundColor:css('--panel-strong'),
        borderColor:css('--border'),
        borderWidth:1,
        titleColor:css('--text'),
        bodyColor:css('--text-soft'),
        callbacks:{label:(item)=>item.dataset.label+': '+item.formattedValue+(yLabel||'')}
      }
    },
    scales:{
      x:{grid:{display:false},ticks:{color:css('--text-faint'),maxTicksLimit:7}},
      y:{grid:{color:'rgba(148,163,184,.12)'},ticks:{color:css('--text-faint')}}
    }
  };
}
function lineDataset(label,color,data){
  return {
    label:label,data:data,borderColor:color,backgroundColor:color+'22',borderWidth:2.2,
    pointRadius:0,pointHoverRadius:4,pointBackgroundColor:color,tension:.36,fill:false
  };
}
function mkLine(id,color,label){return new Chart(el(id),{type:'line',data:{labels:[],datasets:[lineDataset(label,color,[])]},options:chartOptions('')});}
const cLat=mkLine('cLat','#79a8ff','Latency');
const cLoss=mkLine('cLoss','#f6c760','Loss');
const cSpeed=new Chart(el('cSpeed'),{
  type:'line',
  data:{labels:[],datasets:[lineDataset('Up','#b388ff',[]),lineDataset('Down','#33d69f',[])]},
  options:chartOptions(' KB/s')
});
const cScore=mkLine('cScore','#b388ff','Score');
function applyTheme(theme){
  document.body.setAttribute('data-theme',theme);
  localStorage.setItem('zytro-theme',theme);
  el('theme-btn').innerHTML=theme==='light'
    ?'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M21 12.8A9 9 0 1 1 11.2 3 7 7 0 0 0 21 12.8z"/></svg> Light mode'
    :'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M12 3v2m0 14v2M5.64 5.64l1.41 1.41m9.9 9.9 1.41 1.41M3 12h2m14 0h2M5.64 18.36l1.41-1.41m9.9-9.9 1.41-1.41"/><circle cx="12" cy="12" r="4"/></svg> Dark mode';
  refreshChartTheme();
  drawNetMap(latestDevices);
}
function refreshChartTheme(){
  [cLat,cLoss,cSpeed,cScore,modalChart].filter(Boolean).forEach(chart=>{
    chart.options=chartOptions(chart===cSpeed?' KB/s':'');
    chart.update('none');
  });
}
function csrfHeaders(){return {"Content-Type":"application/json","X-Requested-With":"ZytroNet"};}
function openModal(title,labels,datasets,chartType){
  el('modal-title').textContent=title;
  el('modal-overlay').style.display='flex';
  if(modalChart)modalChart.destroy();
  const opts=chartOptions(chartType==='line'?'':'');
  opts.animation={duration:320};
  modalChart=new Chart(el('modal-canvas'),{type:chartType||'line',data:{labels:labels,datasets:datasets},options:opts});
}
function closeModal(){el('modal-overlay').style.display='none';if(modalChart){modalChart.destroy();modalChart=null;}}
document.addEventListener('keydown',(e)=>{if(e.key==='Escape')closeModal();});
function setEventFilter(btn){
  activeFilter=btn.dataset.filter;
  document.querySelectorAll('.event-filter-btn').forEach((b)=>b.classList.remove('active-filter'));
  btn.classList.add('active-filter');
  renderEvents(latestEvents);
}
function renderEvents(evts){
  const rows=activeFilter==='All'?evts:evts.filter((e)=>e.event===activeFilter);
  el('event-count').textContent=rows.length+' entr'+(rows.length===1?'y':'ies');
  if(!rows.length){
    el('evrows').innerHTML='<tr><td colspan="4"><div class="empty-state"><strong>No matching events</strong><span>Try a different filter or wait for more monitoring activity.</span></div></td></tr>';
    return;
  }
  let out='';
  for(const ev of [...rows].reverse()){
    out+=`<tr><td>${escapeHtml(ev.timestamp)}</td><td><span class="${eventBadgeCls(ev.event)}">${escapeHtml(ev.event)}</span></td><td>${escapeHtml(ev.detail||'')}</td><td>${escapeHtml(ev.duration_sec||'—')}</td></tr>`;
  }
  el('evrows').innerHTML=out;
}
function dismissAlert(){alertDismissed=true;el('outage-alert').style.display='none';}
function drawNetMap(devices){
  const cv=el('net-map'),ctx=cv.getContext('2d');
  ctx.clearRect(0,0,cv.width,cv.height);
  ctx.fillStyle=css('--panel-soft');
  ctx.fillRect(0,0,cv.width,cv.height);
  const cx=cv.width/2,cy=cv.height/2;
  nodeHitTargets=[];
  ctx.font='600 13px Inter, Segoe UI, Arial';
  ctx.textAlign='center';
  ctx.fillStyle=css('--accent');
  ctx.beginPath();ctx.arc(cx,cy,20,0,Math.PI*2);ctx.fill();
  ctx.fillStyle=css('--text');
  ctx.fillText('Gateway',cx,cy-30);
  const list=(devices||[]).slice(0,20);
  const ring=list.length>10?132:118;
  list.forEach((d,i)=>{
    const a=(Math.PI*2*i)/Math.max(1,list.length),r=ring,x=cx+Math.cos(a)*r,y=cy+Math.sin(a)*r;
    ctx.strokeStyle='rgba(148,163,184,.24)';
    ctx.lineWidth=1.2;
    ctx.beginPath();ctx.moveTo(cx,cy);ctx.lineTo(x,y);ctx.stroke();
    const online=!!d.alive;
    ctx.fillStyle=online?css('--success'):css('--text-faint');
    ctx.beginPath();ctx.arc(x,y,12,0,Math.PI*2);ctx.fill();
    ctx.fillStyle=css('--text-soft');
    ctx.fillText(d.ip,x,y+28);
    nodeHitTargets.push({ip:d.ip,x,y,r:14});
    if(pendingPings.has(d.ip)){
      ctx.fillStyle=css('--warning');
      ctx.fillText('Pinging...',x,y-20);
    }else if(pingResults[d.ip]&&Date.now()-pingResults[d.ip].at<5000){
      ctx.fillStyle=css('--text');
      ctx.fillText(pingResults[d.ip].label,x,y-20);
    }
  });
}
el('net-map').addEventListener('click',(evt)=>{
  const cv=el('net-map');
  const rect=cv.getBoundingClientRect();
  const scaleX=cv.width/rect.width,scaleY=cv.height/rect.height;
  const x=(evt.clientX-rect.left)*scaleX,y=(evt.clientY-rect.top)*scaleY;
  for(const n of nodeHitTargets){
    const dx=x-n.x,dy=y-n.y;
    if(Math.sqrt(dx*dx+dy*dy)<=n.r){
      if(pendingPings.has(n.ip))return;
      pendingPings.add(n.ip);
      drawNetMap(latestDevices);
      fetch('/ping-device',{method:'POST',headers:csrfHeaders(),body:JSON.stringify({ip:n.ip})})
        .then((r)=>r.json())
        .then((j)=>{pingResults[n.ip]={label:j.error?'Error':(j.latency===null?'Timeout':j.latency.toFixed(1)+' ms'),at:Date.now()};})
        .catch(()=>{pingResults[n.ip]={label:'Error',at:Date.now()};})
        .finally(()=>{pendingPings.delete(n.ip);drawNetMap(latestDevices);});
      return;
    }
  }
});
function setStatusAppearance(status){
  const pill=el('status-pill');
  pill.dataset.status=status||'UNKNOWN';
  pill.textContent=(status||'UNKNOWN').replace('_',' ');
}
function renderTargets(targets){
  el('target-count').textContent=targets.length+' active';
  if(JSON.stringify(targets)!==lastTargetsJson){
    lastTargetsJson=JSON.stringify(targets);
    if(!targets.length){
      el('target-list').innerHTML='<div class="empty-state"><strong>No active targets</strong><span>Add a hostname or IP address to begin monitoring.</span></div>';
      return;
    }
    el('target-list').innerHTML=targets.map((t)=>`
      <div class="target-item">
        <div>
          <div class="target-ip">${escapeHtml(t)}</div>
          <div class="tiny">Active monitoring target</div>
        </div>
        <button class="chip-btn" onclick="removeTarget('${String(t).replace(/'/g,"\\'")}')" type="button">Remove</button>
      </div>
    `).join('');
  }
}
function setPauseUI(paused){
  const pb=el('pause-btn');
  pb.innerHTML=paused
    ?'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="m8 5 11 7-11 7z"/></svg> Resume monitoring'
    :'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M8 5v14M16 5v14"/></svg> Pause monitoring';
  el('pause-banner-btn').textContent=paused?'Resume monitoring':'Pause monitoring';
  el('paused-banner').style.display=paused?'flex':'none';
}
function update(d){
  const n=d.net||{},tr=d.traffic||{},ss=d.session||{},mx=d.metrics||[];
  latestMetrics=mx;latestNet=n;latestEvents=d.events||[];latestDevices=d.devices||[];
  const lastMetric=mx[mx.length-1]||{};
  const labels=mx.map((m)=>m.ts.slice(11,19));
  el('upd').textContent='Updated '+new Date().toLocaleTimeString();
  setStatusAppearance(n.status||'UNKNOWN');
  el('hlat').textContent=n.latency==null?'N/A':n.latency.toFixed(1)+' ms';
  el('hloss').textContent=((n.loss||0).toFixed?n.loss.toFixed(1):n.loss)+'%';
  el('hup').textContent=(ss.up_pct||0).toFixed?(ss.up_pct.toFixed(1)+'%'):'--';
  el('hscore').textContent=(lastMetric.score==null?'--':Number(lastMetric.score).toFixed(1));
  el('hscore-label').textContent=lastMetric.score==null?'Waiting for metrics':scoreLabel(lastMetric.score)+' connection quality';
  el('lip').textContent=d.local_ip||'--';
  el('pip').textContent=d.public_ip||'--';
  el('gw').textContent=d.gateway||'--';
  el('ifc').textContent=d.iface||'--';
  cLat.data.labels=labels;cLat.data.datasets[0].data=mx.map((m)=>m.latency);cLat.update('none');
  cLoss.data.labels=labels;cLoss.data.datasets[0].data=mx.map((m)=>m.loss);cLoss.update('none');
  cSpeed.data.labels=labels;cSpeed.data.datasets[0].data=mx.map((m)=>m.up_bps/1024);cSpeed.data.datasets[1].data=mx.map((m)=>m.dn_bps/1024);cSpeed.update('none');
  cScore.data.labels=labels;cScore.data.datasets[0].data=mx.map((m)=>m.score);cScore.update('none');
  cLat.canvas.onclick=()=>openModal('Latency history',labels,[lineDataset('Latency','#79a8ff',mx.map((m)=>m.latency))],'line');
  cLoss.canvas.onclick=()=>openModal('Packet loss history',labels,[lineDataset('Loss','#f6c760',mx.map((m)=>m.loss))],'line');
  cSpeed.canvas.onclick=()=>openModal('Traffic throughput',labels,[lineDataset('Up','#b388ff',mx.map((m)=>m.up_bps/1024)),lineDataset('Down','#33d69f',mx.map((m)=>m.dn_bps/1024))],'line');
  cScore.canvas.onclick=()=>openModal('Quality score history',labels,[lineDataset('Score','#b388ff',mx.map((m)=>m.score))],'line');
  drawNetMap(latestDevices);
  el('device-count').textContent=(latestDevices.length||0)+' device'+(latestDevices.length===1?'':'s');
  renderEvents(latestEvents);
  if(n.status==='OFFLINE'||n.status==='ISP_FAILURE'){
    if(!alertDismissed){
      el('alert-text').textContent=n.status==='OFFLINE'?'Network outage detected':'ISP failure detected';
      el('outage-alert').style.display='flex';
    }
  }else if(n.status==='ONLINE'||n.status==='DEGRADED'){
    el('outage-alert').style.display='none';
    alertDismissed=false;
  }
  const sbg=el('status-bg');
  if(sbg){
    sbg.style.background=n.status==='OFFLINE'
      ?'radial-gradient(circle at top, rgba(255,107,129,.18), transparent 55%)'
      :n.status==='ISP_FAILURE'
      ?'radial-gradient(circle at top, rgba(255,138,101,.16), transparent 55%)'
      :n.status==='DEGRADED'
      ?'radial-gradient(circle at top, rgba(246,199,96,.14), transparent 55%)'
      :'transparent';
  }
  const paused=!!d.paused;
  setPauseUI(paused);
  const bb=el('beep-btn');
  const be=!!d.beep_enabled;
  bb.className='btn '+(be?'btn-tonal':'btn-ghost');
  bb.innerHTML=be
    ?'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M15 8a5 5 0 0 1 0 8"/><path d="M17.7 5a9 9 0 0 1 0 14"/><path d="M5 10h3l4-4v12l-4-4H5z"/></svg> Alerts sound on'
    :'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M5 10h3l4-4v12l-4-4H5z"/><path d="m19 9-6 6"/><path d="m13 9 6 6"/></svg> Alerts sound off';
  const iv=d.interval||3;
  el('interval-label').textContent=iv.toFixed(1)+'s';
  if(!sliderDragging)el('interval-slider').value=iv;
  if(lastGoodInterval===null)lastGoodInterval=iv;
  renderTargets(d.targets||[]);
  if(!('Notification' in window)){
    el('notif-banner').style.display='none';
  }else{
    if(Notification.permission==='default')el('notif-banner').style.display='flex';
    if(latestEvents.length){
      const last=latestEvents[latestEvents.length-1];
      if(lastEventTs===null){
        lastEventTs=last.timestamp;
      }else if(Notification.permission==='granted'){
        let sent=0;
        for(let i=latestEvents.length-1;i>=0&&sent<3;i--){
          const ev=latestEvents[i];
          if(ev.timestamp<=lastEventTs)break;
          if(ev.event==='OUTAGE_START'||ev.event==='RECOVERY'){
            new Notification('ZytroNet '+ev.event,{body:ev.detail||ev.event});
            sent++;
          }
        }
        lastEventTs=last.timestamp;
      }
    }
  }
}
el('notif-banner').onclick=()=>{
  if(!('Notification' in window))return;
  Notification.requestPermission().then((perm)=>{if(perm!=='default')el('notif-banner').style.display='none';});
};
function addTarget(){
  const t=(el('target-input').value||'').trim();
  fetch('/add-target',{method:'POST',headers:csrfHeaders(),body:JSON.stringify({target:t})})
    .then((r)=>r.json().then((j)=>({ok:r.ok,j})))
    .then((x)=>{
      if(!x.ok){el('target-error').textContent=x.j.error||'Error';return;}
      el('target-input').value='';
      el('target-error').textContent='';
    })
    .catch(()=>{el('target-error').textContent='Request failed';});
}
function removeTarget(ip){
  fetch('/remove-target',{method:'POST',headers:csrfHeaders(),body:JSON.stringify({target:ip})})
    .then((r)=>r.json().then((j)=>({ok:r.ok,j})))
    .then((x)=>{el('target-error').textContent=x.ok?'':(x.j.error||'Error');})
    .catch(()=>{el('target-error').textContent='Request failed';});
}
function togglePause(){
  fetch('/toggle-pause',{method:'POST',headers:{"X-Requested-With":"ZytroNet"}})
    .then((r)=>r.json())
    .then((j)=>{
      setPauseUI(!!j.paused);
      if(window.__lastPayload)window.__lastPayload=Object.assign({},window.__lastPayload,{paused:!!j.paused});
    })
    .catch(()=>{});
}
el('beep-btn').onclick=()=>fetch('/toggle-beep',{method:'POST',headers:{"X-Requested-With":"ZytroNet"}}).then((r)=>r.json()).then((j)=>update(Object.assign({},window.__lastPayload||{}, {beep_enabled:j.beep_enabled}))).catch(()=>{});
el('pause-btn').onclick=togglePause;
el('pause-banner-btn').onclick=togglePause;
el('theme-btn').onclick=()=>applyTheme(document.body.getAttribute('data-theme')==='light'?'dark':'light');
const slider=el('interval-slider');
slider.addEventListener('mousedown',()=>sliderDragging=true);
slider.addEventListener('touchstart',()=>sliderDragging=true);
slider.addEventListener('mouseup',()=>sliderDragging=false);
slider.addEventListener('touchend',()=>sliderDragging=false);
window.addEventListener('mouseup',()=>sliderDragging=false);
window.addEventListener('touchend',()=>sliderDragging=false);
slider.oninput=()=>el('interval-label').textContent=parseFloat(slider.value).toFixed(1)+'s';
slider.onchange=()=>{
  const v=parseFloat(slider.value);
  fetch('/set-interval',{method:'POST',headers:csrfHeaders(),body:JSON.stringify({interval:v})})
    .then((r)=>r.json().then((j)=>({ok:r.ok,j})))
    .then((x)=>{if(x.ok)lastGoodInterval=x.j.interval;else slider.value=lastGoodInterval;})
    .catch(()=>{slider.value=lastGoodInterval;});
};
const savedTheme=localStorage.getItem('zytro-theme');
if(savedTheme==='light'||savedTheme==='dark')applyTheme(savedTheme);
let delay=2000;
function poll(){
  fetch('/data')
    .then((r)=>{if(!r.ok)throw new Error('HTTP '+r.status);return r.json();})
    .then((d)=>{window.__lastPayload=d;update(d);delay=2000;setTimeout(poll,2000);})
    .catch(()=>{el('upd').textContent='Reconnecting in '+(delay/1000).toFixed(0)+'s...';setTimeout(poll,delay);delay=Math.min(delay*1.5,15000);});
}
poll();
</script>
</body>
</html>"""


def build_report_html(state_copy, args, report_metrics=None):
    metrics = report_metrics if report_metrics is not None else state_copy.get("metrics", [])
    if not metrics:
        return None
    events = state_copy.get("events", [])
    ss = state_copy.get("session_summary", {})
    if args.targets is None:
        args.targets = []
    targets = list(state_copy.get("targets", args.targets or []))
    if not targets:
        targets = ["No targets configured"]
    targets_str = ", ".join(targets)

    ts = [m["ts"] for m in metrics]
    lats = [m["latency"] for m in metrics]
    losses = [m["loss"] for m in metrics]
    scores = [m["score"] for m in metrics]
    ups = [round(m["up_bps"] / 1024, 2) for m in metrics]
    dns = [round(m["dn_bps"] / 1024, 2) for m in metrics]
    avg_score = statistics.mean(scores) if scores else 0
    erows = ""
    for ev in events:
        erows += (
            f"<tr><td>{ev['timestamp']}</td><td>{ev['event']}</td><td>{ev['detail']}</td>"
            f"<td>{ev.get('duration_sec','—')}</td></tr>"
        )
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>ZytroNet Report — {datetime.now().strftime('%Y-%m-%d %H:%M')} — Targets: {targets_str}</title>
<meta name="author" content="Elkmirl Kuuku Carter">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root{{--bg:#081120;--panel:#0f1b31;--panel-soft:#14233d;--border:#22314b;--text:#f4f7fb;--text-soft:#b6c3d9;--text-faint:#7f92b1;--accent:#79a8ff;--success:#33d69f;--warning:#f6c760;--danger:#ff6b81;--purple:#b388ff;--shadow:0 24px 60px rgba(0,0,0,.28);}}
*{{box-sizing:border-box}}body{{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:linear-gradient(180deg,#07111f,#0b1628);color:var(--text);padding:28px}}
.shell{{max-width:1280px;margin:0 auto}}
.hero,.panel{{background:rgba(15,27,49,.92);border:1px solid var(--border);border-radius:24px;box-shadow:var(--shadow)}}
.hero{{padding:24px 26px;margin-bottom:18px}}
.eyebrow{{display:inline-block;padding:6px 10px;border-radius:999px;background:rgba(121,168,255,.14);color:var(--accent);font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.04em}}
h1{{margin:14px 0 8px;font-size:30px;letter-spacing:-.03em}}p{{color:var(--text-soft);line-height:1.6}}
.stats{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;margin-top:20px}}
.stat{{padding:16px;border-radius:18px;background:var(--panel-soft);border:1px solid var(--border)}}
.stat .label{{font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:var(--text-faint);font-weight:700}}
.stat .value{{margin-top:10px;font-size:28px;font-weight:800}}
.grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:18px;margin-bottom:18px}}
.panel{{padding:20px}}
.panel h3{{margin:0 0 6px;font-size:17px}}.meta{{color:var(--text-faint);font-size:12px;margin-bottom:14px}}
.chart-wrap{{height:260px}}
table{{width:100%;border-collapse:collapse;border-spacing:0;overflow:hidden}}
th,td{{padding:14px 16px;border-bottom:1px solid var(--border);text-align:left;font-size:13px;vertical-align:top}}
th{{font-size:11px;color:var(--text-faint);text-transform:uppercase;letter-spacing:.06em}}
pre{{margin:0;padding:18px;border-radius:18px;background:#091325;border:1px solid var(--border);overflow:auto;color:var(--text-soft)}}
footer{{margin-top:18px;text-align:center;color:var(--text-faint);font-size:12px}}
@media (max-width:900px){{body{{padding:16px}}.stats,.grid{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<div class="shell">
  <section class="hero">
    <div class="eyebrow">Session report</div>
    <h1>ZytroNet Monitoring Summary</h1>
    <p>Targets: <strong>{targets_str}</strong><br>Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    <div class="stats">
      <div class="stat"><div class="label">Uptime</div><div class="value">{ss.get('up_pct',0):.2f}%</div></div>
      <div class="stat"><div class="label">Runtime</div><div class="value">{fmt_dur(ss.get('total',0))}</div></div>
      <div class="stat"><div class="label">Outages</div><div class="value">{ss.get('outages',0)}</div></div>
      <div class="stat"><div class="label">Avg score</div><div class="value">{avg_score:.1f}</div></div>
    </div>
  </section>

  <section class="grid">
    <article class="panel"><h3>Latency trend</h3><div class="meta">Response time across the captured session.</div><div class="chart-wrap"><canvas id="c1"></canvas></div></article>
    <article class="panel"><h3>Packet loss</h3><div class="meta">Loss percentage over time.</div><div class="chart-wrap"><canvas id="c2"></canvas></div></article>
    <article class="panel"><h3>Quality score</h3><div class="meta">Composite health score from the monitoring run.</div><div class="chart-wrap"><canvas id="c3"></canvas></div></article>
    <article class="panel"><h3>Traffic throughput</h3><div class="meta">Upload and download volume in KB/s.</div><div class="chart-wrap"><canvas id="c4"></canvas></div></article>
  </section>

  <section class="panel" style="margin-bottom:18px">
    <h3>Events</h3>
    <div class="meta">Operational activity captured during the session.</div>
    <table>
      <thead><tr><th>Timestamp</th><th>Event</th><th>Detail</th><th>Duration</th></tr></thead>
      <tbody>{erows or "<tr><td colspan='4'>No events</td></tr>"}</tbody>
    </table>
  </section>

  <section class="panel">
    <h3>Per-target snapshot</h3>
    <div class="meta">Per-target data reflects the last completed monitoring cycle only.</div>
    <pre>{json.dumps(state_copy.get("net",{}).get("target_results",{}), indent=2)}</pre>
  </section>

  <footer>ZytroNet report export · Engineered by Elkmirl Kuuku Carter</footer>
</div>
<script>
const ts={json.dumps(ts)},lats={json.dumps(lats)},losses={json.dumps(losses)},scores={json.dumps(scores)},ups={json.dumps(ups)},dns={json.dumps(dns)};
const baseOptions={{
  responsive:true,
  maintainAspectRatio:false,
  interaction:{{mode:'index',intersect:false}},
  plugins:{{legend:{{labels:{{color:'#b6c3d9',boxWidth:10,boxHeight:10,usePointStyle:true}}}}}},
  scales:{{x:{{grid:{{display:false}},ticks:{{color:'#7f92b1'}}}},y:{{grid:{{color:'rgba(148,163,184,.12)'}},ticks:{{color:'#7f92b1'}}}}}}
}};
function ds(label,data,color){{return {{label:label,data:data,borderColor:color,pointRadius:0,borderWidth:2.2,tension:.36}};}}
new Chart(document.getElementById('c1'),{{type:'line',data:{{labels:ts,datasets:[ds('Latency',lats,'#79a8ff')]}},options:baseOptions}});
new Chart(document.getElementById('c2'),{{type:'line',data:{{labels:ts,datasets:[ds('Loss',losses,'#f6c760')]}},options:baseOptions}});
new Chart(document.getElementById('c3'),{{type:'line',data:{{labels:ts,datasets:[ds('Score',scores,'#b388ff')]}},options:baseOptions}});
new Chart(document.getElementById('c4'),{{type:'line',data:{{labels:ts,datasets:[ds('Up',ups,'#b388ff'),ds('Down',dns,'#33d69f')]}},options:baseOptions}});
</script>
</body>
</html>"""


def generate_report(state_copy, args, report_metrics=None):
    html = build_report_html(state_copy, args, report_metrics)
    if html is None:
        return False
    try:
        with open(args.report, "w", encoding="utf-8") as f:
            f.write(html)
        return True
    except Exception as e:
        print(f"Report error: {e}")
        return False


def generate_excel_report(state_copy, args, report_metrics):
    if not OPENPYXL_OK:
        return False
    metrics = report_metrics
    if not metrics:
        return False
    events = state_copy.get("events", [])
    ss = state_copy.get("session_summary", {})
    output_path = os.path.splitext(args.report)[0] + ".xlsx"
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Summary")
        items = [
            ("Runtime", fmt_dur(ss.get("total", 0))),
            ("Uptime", fmt_dur(ss.get("uptime", 0))),
            ("Downtime", fmt_dur(ss.get("downtime", 0))),
            ("Uptime %", f"{ss.get('up_pct',0):.2f}%"),
            ("Outages", ss.get("outages", 0)),
            ("MTBO", fmt_dur(ss.get("mtbo")) if ss.get("mtbo") is not None and ss.get("mtbo") != 0 else ("0s" if ss.get("mtbo") == 0 else "N/A")),
        ]
        for i, (k, v) in enumerate(items, 1):
            ws.cell(i, 1, k).font = Font(bold=True)
            ws.cell(i, 2, v).alignment = Alignment(horizontal="right")

        wm = wb.create_sheet("Metrics History")
        hdr = ["TS", "Status", "Latency", "Loss", "Jitter", "Score", "Up_bps", "Dn_bps", "DNS", "GW"]
        wm.append(hdr)
        for m in metrics:
            wm.append([m["ts"], m["status"], m["latency"], m["loss"], m["jitter"], m["score"], m["up_bps"], m["dn_bps"], m["dns_ok"], m["gw_ok"]])
        wm.conditional_formatting.add(f"B2:B{len(metrics)+1}", CellIsRule(operator="equal", formula=['"ONLINE"'], fill=PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")))
        for r in range(2, len(metrics) + 2):
            v = wm.cell(r, 2).value
            fill = "00AA00" if v == "ONLINE" else "CC0000" if v == "OFFLINE" else "AA8800"
            wm.cell(r, 2).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            wm.cell(r, 2).font = Font(color="FFFFFF", bold=True)
        lc = LineChart(); lc.title = "Latency"; lc.add_data(Reference(wm, min_col=3, min_row=1, max_row=len(metrics)+1), titles_from_data=True); lc.set_categories(Reference(wm, min_col=1, min_row=2, max_row=len(metrics)+1)); wm.add_chart(lc, "L2")
        sc = LineChart(); sc.title = "Score"; sc.add_data(Reference(wm, min_col=6, min_row=1, max_row=len(metrics)+1), titles_from_data=True); sc.set_categories(Reference(wm, min_col=1, min_row=2, max_row=len(metrics)+1)); wm.add_chart(sc, "L18")

        we = wb.create_sheet("Event Log")
        we.append(["Timestamp", "Event", "Detail", "Duration"])
        for ev in events:
            we.append([ev.get("timestamp", ""), ev.get("event", ""), ev.get("detail", ""), ev.get("duration_sec", "")])
        for r in range(2, len(events) + 2):
            ev_type = we.cell(r, 2).value
            fill = "CC0000" if ev_type == "OUTAGE_START" else "00AA00" if ev_type == "RECOVERY" else "AA8800" if ev_type == "STATUS_CHANGE" else "0066AA"
            we.cell(r, 2).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            we.cell(r, 2).font = Font(color="FFFFFF", bold=True)

        wt = wb.create_sheet("Per-Target Results")
        wt.append(["Target", "Latency", "Loss", "Jitter"])
        for t, d in (state_copy.get("net", {}).get("target_results", {}) or {}).items():
            wt.append([t, "Timeout" if d.get("latency") is None else d.get("latency"), d.get("loss"), d.get("jitter")])

        for sheet in [ws, wm, we, wt]:
            for col in range(1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(col)].width = 18
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Excel report error: {e}")
        return False


# ── Flask App ─────────────────────────────────────────────────────────────────
def build_flask_app(state, pauser, web_password=None, args=None):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    web_root = os.path.join(base_dir, "zytronet_web")
    template_dir = os.path.join(web_root, "templates")
    static_dir = os.path.join(web_root, "static")
    app = Flask(
        __name__,
        template_folder=template_dir,
        static_folder=static_dir,
        static_url_path="/static",
    )
    import logging
    logging.getLogger("werkzeug").setLevel(logging.ERROR)
    app.logger.disabled = True

    # HTTP Basic Auth transmits credentials unencrypted. Suitable for trusted LAN use only. Use ZYTRO_WEB_PASSWORD env var to avoid exposing credentials in the process table.
    def check_auth(request_obj):
        if web_password is None:
            return True
        auth = request_obj.headers.get("Authorization", "")
        if not auth.startswith("Basic "):
            return False
        try:
            decoded = base64.b64decode(auth.split(" ", 1)[1]).decode("utf-8")
        except Exception:
            return False
        _, sep, pw = decoded.partition(":")
        if sep == "":
            return False
        return hmac.compare_digest(pw, web_password)

    def auth_required_response():
        return Response("Unauthorised", status=401, headers={"WWW-Authenticate": 'Basic realm="ZytroNet"'})

    def check_csrf(request_obj):
        return request_obj.headers.get("X-Requested-With") == "ZytroNet"

    def csrf_fail():
        return Response(json.dumps({"error": "CSRF check failed"}), status=403, mimetype="application/json")

    @app.route("/")
    def index():
        if not check_auth(__import__("flask").request):
            return auth_required_response()
        try:
            return render_template("index.html")
        except Exception:
            # Safe fallback: keep single-file distribution working if templates are missing.
            return render_template_string(WEB_HTML)

    @app.route("/data")
    def data():
        request = __import__("flask").request
        if not check_auth(request):
            return auth_required_response()
        with state["targets_lock"]:
            targets_copy = list(state.get("targets") or [])
        with state["lock"]:
            payload = {
                "net": state.get("net") or {},
                "traffic": state.get("traffic") or {"up": 0, "dn": 0, "sent": 0, "recv": 0},
                "session": state.get("session_summary") or {},
                "devices": state.get("devices") or [],
                "events": state.get("events") or [],
                "metrics": state.get("metrics") or [],
                "local_ip": state.get("local_ip", "--"),
                "public_ip": state.get("public_ip", "--"),
                "gateway": state.get("gateway", "--"),
                "iface": state.get("iface", "--"),
                "interval": state.get("interval", 3.0),
                "beep_enabled": state.get("beep_enabled", True),
                "paused": pauser.is_paused(),
                "targets": targets_copy,
            }
        return Response(json.dumps(payload), mimetype="application/json")

    @app.route("/toggle-beep", methods=["POST"])
    def toggle_beep():
        request = __import__("flask").request
        if not check_auth(request):
            return auth_required_response()
        if not check_csrf(request):
            return csrf_fail()
        global BEEP_ENABLED
        with _beep_lock:
            BEEP_ENABLED = not BEEP_ENABLED
            new_val = BEEP_ENABLED
        with state["lock"]:
            state["beep_enabled"] = new_val
        return Response(json.dumps({"beep_enabled": new_val}), mimetype="application/json")

    @app.route("/toggle-pause", methods=["POST"])
    def toggle_pause():
        request = __import__("flask").request
        if not check_auth(request):
            return auth_required_response()
        if not check_csrf(request):
            return csrf_fail()
        new_val = pauser.toggle()
        return Response(json.dumps({"paused": new_val}), mimetype="application/json")

    @app.route("/add-target", methods=["POST"])
    def add_target():
        request = __import__("flask").request
        if not check_auth(request):
            return auth_required_response()
        if not check_csrf(request):
            return csrf_fail()
        data = request.get_json(force=False, silent=True)
        if data is None:
            return Response(json.dumps({"error": "Invalid JSON"}), status=400, mimetype="application/json")
        value = str(data.get("target", "")).strip()
        if not re.match(r"^[a-zA-Z0-9.\-]{1,253}$", value):
            return Response(json.dumps({"error": "Invalid target format"}), status=400, mimetype="application/json")
        with state["targets_lock"]:
            if value in state["targets"]:
                return Response(json.dumps({"error": "Target already exists"}), status=409, mimetype="application/json")
            state["targets"].append(value)
            cur = list(state["targets"])
        return Response(json.dumps({"targets": cur}), mimetype="application/json")

    @app.route("/remove-target", methods=["POST"])
    def remove_target():
        request = __import__("flask").request
        if not check_auth(request):
            return auth_required_response()
        if not check_csrf(request):
            return csrf_fail()
        data = request.get_json(force=False, silent=True)
        if data is None:
            return Response(json.dumps({"error": "Invalid JSON"}), status=400, mimetype="application/json")
        value = str(data.get("target", "")).strip()
        if not re.match(r"^[a-zA-Z0-9.\-]{1,253}$", value):
            return Response(json.dumps({"error": "Invalid target format"}), status=400, mimetype="application/json")
        with state["targets_lock"]:
            if len(state["targets"]) == 1 and state["targets"][0] == value:
                return Response(json.dumps({"error": "Cannot remove the last target"}), status=400, mimetype="application/json")
            if value not in state["targets"]:
                return Response(json.dumps({"targets": list(state["targets"])}), mimetype="application/json")
            state["targets"].remove(value)
            cur = list(state["targets"])
        return Response(json.dumps({"targets": cur}), mimetype="application/json")

    @app.route("/set-interval", methods=["POST"])
    def set_interval():
        request = __import__("flask").request
        if not check_auth(request):
            return auth_required_response()
        if not check_csrf(request):
            return csrf_fail()
        data = request.get_json(force=False, silent=True)
        if data is None:
            return Response(json.dumps({"error": "Invalid JSON"}), status=400, mimetype="application/json")
        value = data.get("interval")
        if not isinstance(value, (int, float)):
            return Response(json.dumps({"error": "Interval must be numeric"}), status=400, mimetype="application/json")
        fv = float(value)
        if fv < 0.5 or fv > 60.0:
            return Response(json.dumps({"error": "Interval out of range"}), status=400, mimetype="application/json")
        with state["lock"]:
            state["interval"] = fv
        return Response(json.dumps({"interval": fv}), mimetype="application/json")

    @app.route("/report")
    def report_route():
        request = __import__("flask").request
        if not check_auth(request):
            return auth_required_response()
        with state["lock"]:
            state_copy = {
                "net": state.get("net") or {},
                "traffic": state.get("traffic") or {},
                "session_summary": state.get("session_summary") or {},
                "devices": list(state.get("devices") or []),
                "events": list(state.get("events") or []),
                "metrics": list(state.get("metrics") or []),
                "local_ip": state.get("local_ip", "--"),
                "public_ip": state.get("public_ip", "--"),
                "gateway": state.get("gateway", "--"),
                "iface": state.get("iface", "--"),
                "interval": state.get("interval", 3.0),
                "beep_enabled": state.get("beep_enabled", True),
            }
        with state["targets_lock"]:
            state_copy["targets"] = list(state.get("targets") or [])
        try:
            html = build_report_html(state_copy, args)
            if html is None:
                return Response("Not enough data yet.", status=503)
            return Response(html, mimetype="text/html", headers={"Content-Disposition": 'attachment; filename="zytronet_report.html"'})
        except Exception as e:
            return Response(f"Report generation error: {e}", status=500)

    @app.route("/ping-device", methods=["POST"])
    def ping_device():
        request = __import__("flask").request
        if not check_auth(request):
            return auth_required_response()
        if not check_csrf(request):
            return csrf_fail()
        data = request.get_json(force=False, silent=True)
        if data is None:
            return Response(json.dumps({"error": "Invalid JSON"}), status=400, mimetype="application/json")
        ip = str(data.get("ip", "")).strip()
        if not re.match(r"^(\d{1,3}\.){3}\d{1,3}$", ip):
            return Response(json.dumps({"error": "Invalid IP"}), status=400, mimetype="application/json")
        try:
            octets = [int(x) for x in ip.split(".")]
            if any(x < 0 or x > 255 for x in octets):
                raise ValueError()
        except Exception:
            return Response(json.dumps({"error": "Invalid IP"}), status=400, mimetype="application/json")
        # state["devices"] is populated by ARP scan, which is an unauthenticated network protocol.
        with state["lock"]:
            known = [d["ip"] for d in state.get("devices", []) if isinstance(d, dict) and "ip" in d]
        if ip not in known:
            return Response(json.dumps({"error": "Device not in discovered list"}), status=403, mimetype="application/json")
        with _ping_in_flight_lock:
            if ip in _ping_in_flight:
                return Response(json.dumps({"error": "Ping already in progress for this IP"}), status=429, mimetype="application/json")
            _ping_in_flight.add(ip)
        try:
            future = _ping_executor.submit(ping_host, ip, 3)
            lat, loss, jitter = future.result(timeout=6)
        except concurrent.futures.TimeoutError:
            with _ping_in_flight_lock:
                _ping_in_flight.discard(ip)
            return Response(json.dumps({"error": "Ping timeout"}), status=504, mimetype="application/json")
        except Exception as e:
            with _ping_in_flight_lock:
                _ping_in_flight.discard(ip)
            return Response(json.dumps({"error": str(e)}), status=500, mimetype="application/json")
        with _ping_in_flight_lock:
            _ping_in_flight.discard(ip)
        return Response(json.dumps({"ip": ip, "latency": lat, "loss": loss, "jitter": jitter}), mimetype="application/json")

    return app


# ── Entry Point ───────────────────────────────────────────────────────────────
def main():
    global NO_COLOR
    args = parse_args()
    if args.web_password is not None and not args.web_password.strip():
        args.web_password = None
    if args.web_password is None:
        args.web_password = os.environ.get("ZYTRO_WEB_PASSWORD") or None
    if args.no_color: NO_COLOR = True
    if IS_WIN: os.system("")  # enable ANSI on Windows

    _shutdown_event.clear()

    print(cl("\n  ◈ ZytroNet Monitoring Engine v3.0", A.CYN, A.BOLD))
    print(cl("  Initialising...\n", A.GRY))

    # Warnings
    no_priv = False
    if not args.no_arp and not is_root():
        print(cl("  ⚠  Run as admin/sudo for full ARP device discovery.\n"
                 "     Continuing with limited scan capability.\n", A.YLW))
        no_priv = True

    if not PSUTIL_OK:
        print(cl("  ⚠  psutil not found — traffic monitoring disabled.\n"
                 "     Install: python -m pip install psutil\n", A.YLW))

    if not FLASK_OK and not args.no_web:
        print(cl("  ⚠  Flask not found — web dashboard disabled.\n"
                 "     Install: python -m pip install flask\n", A.YLW))
    if not OPENPYXL_OK:
        print(cl("  ⚠  openpyxl not found — Excel report disabled. Install: pip install openpyxl", A.YLW))

    # Detect network info
    iface, local_ip = get_primary_interface()
    gateway         = get_gateway()
    subnet          = get_local_subnet(local_ip)

    print(cl(f"  Interface : {iface or 'Unknown'}", A.GRY))
    print(cl(f"  Local IP  : {local_ip}", A.GRY))
    print(cl(f"  Subnet    : {subnet or 'Unknown'}", A.GRY))
    print(cl(f"  Gateway   : {gateway}", A.GRY))
    print(cl(f"  Targets   : {', '.join(args.targets)}", A.GRY))
    print(cl("  Public IP : fetching...", A.GRY), end="", flush=True)
    public_ip = get_public_ip()
    print(cl(f"\r  Public IP : {public_ip}", A.MGT))

    if not args.no_web and FLASK_OK:
        print(cl(f"  Web dash  : http://localhost:{args.web_port}  "
                 f"(also accessible on your network)", A.CYN))
        print(cl("  ⚠  Web server bound to 0.0.0.0 — accessible on all interfaces. Use --web-password and consider firewall rules.", A.YLW))
        if args.web_password:
            print(cl("  ⚠  Web server bound to 0.0.0.0 — credentials transmitted unencrypted. Ensure you are on a trusted network or use a firewall to restrict access.", A.YLW))
            print(cl("  Tip: set ZYTRO_WEB_PASSWORD env var instead of --web-password to avoid password in process table.", A.GRY))
            print(cl("  Auth      : enabled", A.GRY))
        else:
            print(cl("  Auth      : disabled", A.GRY))
    print(cl(f"  Report    : {args.report}\n", A.GRY))
    print(cl("  Starting in 2 seconds...", A.GRY))
    time.sleep(2)

    # Shared state dict for web server
    active_targets = list(args.targets)
    state = {
        "lock": threading.Lock(),
        "targets_lock": threading.Lock(),
        "net": {}, "traffic": {"up":0,"dn":0,"sent":0,"recv":0},
        "session_summary": {}, "devices": [], "events": [], "metrics": [],
        "local_ip": local_ip, "public_ip": public_ip,
        "gateway": gateway, "iface": iface,
        "beep_enabled": True,
        "interval": args.interval,
        "targets": active_targets,
    }
    state_ready = threading.Event()

    # Initialise components
    event_log   = EventLog(maxlen=500, max_age=args.max_age)
    metrics_st  = MetricsStore(maxlen=args.history, max_age=args.max_age)
    report_metrics_st = MetricsStore(maxlen=99999, max_age=86400)
    session     = Session()
    traffic_mon = TrafficMonitor(iface=iface)
    scanner     = ARPScanner(enabled=not args.no_arp, local_ip=local_ip)
    lat_hist    = deque(maxlen=60)
    dashboard   = TerminalDashboard()
    pauser      = PauseController(state)

    if not args.no_arp:
        scanner.start(interval=args.arp_interval)

    start_interface_watcher(state, event_log, scanner)

    # State startup barrier before Flask thread starts
    state_ready.set()

    # Start Flask web server
    if not args.no_web and FLASK_OK:
        flask_app = build_flask_app(state, pauser, web_password=args.web_password, args=args)
        def start_flask():
            try:
                flask_app.run(
                    host="0.0.0.0", port=args.web_port,
                    debug=False, use_reloader=False, threaded=True)
            except Exception as e:
                print(cl(f"\n  ⚠  Web dashboard failed to start: {e}", A.RED))
        threading.Thread(target=start_flask, daemon=True).start()
        time.sleep(0.5)  # Give Flask time to start and potentially fail

    # Start pause listener
    pauser.start()

    # Clear terminal
    sys.stdout.write(A.CLR + A.HIDE)
    sys.stdout.flush()

    # ── Main Loop ─────────────────────────────────────────────────────────────
    try:
        while not _shutdown_event.is_set():
            if pauser.is_paused():
                _shutdown_event.wait(0.2)
                continue

            with state["targets_lock"]:
                targets_snapshot = list(state["targets"])
            with state["lock"]:
                gateway_for_cycle = state.get("gateway", gateway)
            net = get_network_status(gateway_for_cycle, targets_snapshot, args.ping_count)
            if net["latency"]:
                lat_hist.append(net["latency"])

            traffic_mon.update()
            tr = traffic_mon.snap()

            session.update(net["status"], event_log)
            ss = session.summary()

            metrics_st.add(net, tr)
            report_metrics_st.add(net, tr)

            devices, last_scan = scanner.get()
            events_snapshot = event_log.all()
            metrics_snapshot = metrics_st.all_locked()
            with _beep_lock:
                beep_enabled_snapshot = BEEP_ENABLED
            with state["lock"]:
                public_ip_snapshot = state["public_ip"]
                interval_snapshot = state["interval"]
                local_ip_snapshot = state["local_ip"]
                gateway_snapshot = state["gateway"]
                iface_snapshot = state["iface"]

            # Sync shared state
            with state["lock"]:
                state["net"]             = net
                state["traffic"]         = tr
                state["session_summary"] = ss
                state["devices"]         = devices
                state["events"]          = events_snapshot
                state["metrics"]         = metrics_snapshot
                state["beep_enabled"]    = beep_enabled_snapshot

            # Draw terminal dashboard
            dashboard.draw(
                net, tr, devices, last_scan,
                local_ip_snapshot, public_ip_snapshot, gateway_snapshot, iface_snapshot,
                session, event_log, metrics_st,
                lat_hist, pauser.is_paused(), no_priv, args, beep_enabled_snapshot, interval_snapshot
            )

            _shutdown_event.wait(interval_snapshot)

    except KeyboardInterrupt:
        _shutdown_event.set()
        pauser.stop()
        scanner.stop()
        sys.stdout.write(A.SHOW + "\n")
        sys.stdout.flush()

        time.sleep(0.3)

        report_snapshot = report_metrics_st.all_locked()
        with state["lock"]:
            state["session_summary"] = session.summary()
            state_copy = {
                "net": state.get("net") or {},
                "traffic": state.get("traffic") or {"up":0,"dn":0,"sent":0,"recv":0},
                "session_summary": state.get("session_summary") or {},
                "devices": list(state.get("devices") or []),
                "events": list(state.get("events") or []),
                "metrics": list(report_snapshot),
                "local_ip": state.get("local_ip", "--"),
                "public_ip": state.get("public_ip", "--"),
                "gateway": state.get("gateway", "--"),
                "iface": state.get("iface", "--"),
                "beep_enabled": state.get("beep_enabled", True),
                "interval": state.get("interval", args.interval),
            }
        with state["targets_lock"]:
            state_copy["targets"] = list(state["targets"])

        ss = session.summary()
        print(cl("\n  ◈ ZytroNet — Session Summary", A.CYN, A.BOLD))
        print(cl("  " + "═" * 50, A.CYN))

        def row(l, v, c=A.WHT):
            print(f"  {cl(l+':', A.GRY):<32} {cl(v, c)}")

        row("Total Runtime",     fmt_dur(ss["total"]))
        row("Uptime",            fmt_dur(ss["uptime"]),   A.GRN)
        row("Downtime",          fmt_dur(ss["downtime"]),
            A.RED if ss["downtime"] > 0 else A.GRY)
        row("Uptime %",          f"{ss['up_pct']:.2f}%",
            A.GRN if ss["up_pct"] > 95 else A.YLW)
        row("Total Outages",     str(ss["outages"]),
            A.RED if ss["outages"] > 0 else A.GRN)
        if ss["mtbo"]:
            row("Avg Outage Duration", fmt_dur(ss["mtbo"]), A.YLW)

        print(cl("\n  Generating HTML report...", A.GRY))
        ok = generate_report(state_copy, args, report_metrics=report_snapshot)
        if ok:
            print(cl(f"  Saved: {args.report}", A.GRN))
            print(cl("  Open in your browser to view all charts.\n", A.GRY))
        else:
            print(cl("  Not enough data to generate report.\n", A.YLW))

        if generate_excel_report(state_copy, args, report_metrics=report_snapshot):
            print(cl(f"  Saved: {os.path.splitext(args.report)[0] + '.xlsx'}", A.GRN))
        elif OPENPYXL_OK:
            print(cl("  Excel report skipped (not enough data).", A.YLW))

        print(cl("  Thank you for using ZytroNet.\n", A.CYN))


if __name__ == "__main__":
    main()
